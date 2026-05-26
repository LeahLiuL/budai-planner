from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

pkg_col = col_map.get('PKG', 21)
pkge_col = col_map.get('PKG EB', 29)
teu_col = col_map.get('EFF. TEUS', 12)
vessel_col = col_map.get('CUL VESSELS', 3)
week_col = col_map.get('WEEK', 2)
service_col = col_map.get('SERVICES', 5)
voyage_col = col_map.get('VOYAGE .No', 4)
remarks_col = col_map.get('REMARKS', 34)

dest_cols = {}
for pn in ['SOK','JED','MUN','NGB']:
    if pn in col_map: dest_cols[pn] = col_map[pn]

def fmt(d):
    if d is None: return ''
    if isinstance(d, datetime): return d.strftime('%Y-%m-%d')
    s=str(d)[:10]
    return s if s else ''

def parse_dt(s):
    if not s: return None
    try: return datetime.strptime(str(s)[:10], '%Y-%m-%d')
    except:
        try: return datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
        except: return None

def get_rgb(cell):
    try: return str(cell.fill.fgColor.rgb).upper() if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb else '00000000'
    except: return '00000000'

def classify(rgb):
    if rgb=='00000000': return 'WHITE'
    if 'DAF2D0' in rgb or 'DAF' in rgb: return 'PINK'
    if any(x in rgb for x in ['FBE2D5','CAEDFB']): return 'FEEDER'
    if rgb.startswith('FF') and len(rgb)==8 and rgb[2:]!='000000': return 'OTHER'
    return 'UNKNOWN'

# Collect R51-R77
ships = []
for row in range(51, 80):
    vc=sheet.cell(row=row,column=vessel_col)
    v=str(vc.value or '').strip()
    if not v: continue
    
    rgb=get_rgb(vc); ctype=classify(rgb)
    
    pkg=sheet.cell(row=row,column=pkg_col).value; teu_val=sheet.cell(row=row,column=teu_col).value
    wk=sheet.cell(row=row,column=week_col).value; svc=sheet.cell(row=row,column=service_col).value
    voy=sheet.cell(row, column=voyage_col).value if voyage_col else None
    rmk=sheet.cell(row=row,column=remarks_col).value
    
    # Parse TEU as int
    try: teu=int(float(teu_val)) if teu_val else 0
    except: teu=0
    
    dst={}
    for pn,pc in dest_cols.items():
        val=sheet.cell(row=row,column=pc).value
        if val: dst[pn]=fmt(val)
    
    ships.append({
        'row':row,'vessel':v,'week':wk,'svc':str(svc) if svc else '',
        'voy':voy,'pkg':fmt(pkg),'teu':teu,'rgb':rgb,'type':ctype,
        'rmk':(str(rmk)[:60] if rmk else ''),'dst':dst,
        'pkg_dt':parse_dt(pkg)
    })

white=[s for s in ships if s['type']=='WHITE']
feeders=[s for s in ships if s['type']=='FEEDER']
pink=[s for s in ships if s['type']=='PINK']

# ============================================================
# OPTIMIZATION ALGORITHM: Max BSA Utilization + Min Wait Time
# ============================================================
# Strategy: Weighted score = -alpha*wait_days + beta*fill_rate_bonus
# alpha = weight for time (higher = more important to minimize wait)
# beta  = weight for utilization (higher = more important to fill ships)
# We use a greedy approach that scores each (mainline, feeder) assignment

ALPHA = 10.0   # wait day penalty (per day)
BETA  = 15.0   # fill rate bonus (proportional to how full the feeder gets)

MIN_WAIT_DAYS = 1  # minimum operation window at PKG

# Build all valid assignments with scores
assignments = []  # list of dicts
for mw in white:
    eta = mw['pkg_dt']
    if not eta or mw['teu'] <= 0:
        continue
    cargo = mw['teu']  # assume ship is full = cargo volume
    
    for mf in feeders:
        etd = mf['pkg_dt']
        if not etd:
            continue
        
        delta = (etd - eta).days
        if delta < MIN_WAIT_DAYS:
            continue
        
        # Score: lower wait = better, higher fill potential = better
        # fill_potential: if this cargo goes here, what % of feeder capacity would be used?
        bsa = mf['teu']
        if bsa <= 0:
            continue
        
        fill_after = min(cargo, bsa) / bsa  # fraction of BSA used by this assignment alone
        
        # Combined score: we want HIGH score = good assignment
        score = -ALPHA * delta + BETA * fill_after * 100
        
        assignments.append({
            'mainline': mw,
            'feeder': mf,
            'wait_days': delta,
            'cargo_teu': cargo,
            'bsa_capacity': bsa,
            'score': score,
            'assigned_teu': min(cargo, bsa),
        })

# Sort by score descending (best first)
assignments.sort(key=lambda x: x['score'], reverse=True)

# Greedy assignment with tracking
feeder_remaining = {f['vessel']+str(f['row']): f['teu'] for f in feeders}
feeder_assigned = defaultdict(list)  # feeder_key -> list of assigned cargos
mainline_assigned = defaultdict(list)  # mainline_key -> list of assigned feeders

final_assignments = []
total_assigned_teu = 0
total_cargo_teu = sum(m['teu'] for m in white if m['teu'])

# Track which mainline vessels still have unassigned cargo
mainline_remaining = {}
for m in white:
    mainline_remaining[m['vessel']+str(m['row'])] = m['teu']

for a in assignments:
    mk = a['mainline']['vessel']+str(a['mainline']['row'])
    fk = a['feeder']['vessel']+str(a['feeder']['row'])
    
    rem_main = mainline_remaining.get(mk, 0)
    rem_feed = feeder_remaining.get(fk, 0)
    
    if rem_main <= 0 or rem_feed <= 0:
        continue
    
    assign_teu = min(rem_main, rem_feed, a['cargo_teu'])
    
    final_assignments.append({
        'mainline': a['mainline'],
        'feeder': a['feeder'],
        'wait_days': a['wait_days'],
        'assigned_teu': assign_teu,
        'bsa_cap': a['bsa_capacity'],
        'score': a['score'],
    })
    
    mainline_remaining[mk] -= assign_teu
    feeder_remaining[fk] -= assign_teu
    total_assigned_teu += assign_teu

# Compute statistics
total_bsa = sum(f['teu'] for f in feeders if f['teu'])
used_bsa = total_bsa - sum(feeder_remaining.values())
util_pct = (used_bsa / total_bsa * 100) if total_bsa > 0 else 0
avg_wait = sum(a['wait_days']*a['assigned_teu'] for a in final_assignments) / total_assigned_teu if total_assigned_teu > 0 else 0

print("=== Optimization Results ===")
print(f"White(main): {len(white)}, Feeders: {len(feeders)}, Pink(direct): {len(pink)}")
print(f"Total cargo: {total_cargo_teu} TEU")
print(f"Assigned: {total_assigned_teu} TEU ({total_assigned_teu/total_cargo_teu*100:.1f}% of cargo)")
print(f"BSA Utilized: {used_bsa}/{total_bsa} TEU ({util_pct:.1f}%)")
print(f"Weighted Avg Wait: {avg_wait:.1f} days")
print(f"Total assignments: {len(final_assignments)}")

# ============================================================
# BUILD HTML REPORT
# ============================================================
L=[]
A=L.append

A('<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">')
A('<title>布袋船优化排计划 - BSA利用率最大化 + 中转时间最短</title>')
A('''<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:#f0f2f5;padding:20px;color:#333}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:18px;overflow:hidden}
.card-head{padding:14px 18px;font-weight:bold;font-size:16px;display:flex;align-items:center;gap:10px;border-bottom:1px solid #eee}
.card-sub{padding:10px 18px;font-size:13px;color:#555;border-bottom:1px solid #f0f0f0;background:#fafbfc}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f5f6fa;padding:8px 10px;text-align:left;font-weight:600;border-bottom:2px solid #dde1e6;position:sticky;top:0}
td{padding:6px 10px;border-bottom:1px solid #f0f0f0;white-space:nowrap}
tr:hover td{background:#f8f9ff}
.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:normal}
.bg-w{background:#e8eaed;color:#444}.bg-p{background:#fce4ec;color:#c2185b}.bg-f{background:#fff3e0;color:#e65100}
.date-f{font-family:"Courier New",monospace;font-size:12px}
.teu-n{font-weight:600;text-align:right}
.good{color:#2e7d32;font-weight:600}.warn{color:#e65100;font-weight:600}.bad{color:#c62828;font-weight:600}
.rmks{color:#888;font-style:italic;font-size:11px;max-width:220px}

/* KPI cards */
.kpi-grid{display:flex;gap:16px;padding:18px;flex-wrap:wrap}
.kpi{min-width:140px;text-align:center;padding:12px;border-radius:8px;background:#f8f9ff}
.kpi-val{font-size:28px;font-weight:bold;color:#1565c0}
.kpi-lbl{font-size:11px;color:#666;margin-top:4px}
.kpi-green .kpi-val{color:#2e7d32}
.kpi-orange .kpi-val{color:#e65100}

/* Assignment card */
.asgn-card{border:1px solid #e0e0e0;border-radius:8px;margin-bottom:12px;overflow:hidden}
.asgn-head{padding:10px 14px;display:flex;align-items:center;gap:12px;font-weight:bold;font-size:13px;
           background:linear-gradient(135deg,#e3f2fd,#f3e5f5)}
.asgn-body{padding:10px 14px}
.asgn-row{display:flex;align-items:center;padding:5px 0;border-bottom:1px dotted #e8e8e8;font-size:12px}
.asgn-row:last-child{border:none}

/* Progress bar */
.progress-wrap{width:120px;height:18px;background:#e8e8e8;border-radius:9px;overflow:hidden;display:inline-block;vertical-align:middle}
.progress-fill{height:100%;border-radius:9px;transition:width .3s}
.progress-fill.green{background:linear-gradient(90deg,#81c784,#4caf50)}
.progress-fill.yellow{background:linear-gradient(90deg,#fff176,#ffc107)}
.progress-fill.red{background:linear-gradient(90deg,#ef9a9a,#f44336)}

.legend-bar{padding:10px 18px;display:flex;gap:24px;font-size:12px;border-top:1px solid #eee;background:#fafafa}
.color-dot{width:20px;height:20px;border-radius:4px;border:1px solid #ccc;display:inline-block;vertical-align:middle}

.score-tag{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
</style></head><body>
''')

# Header KPIs
A('<div class="card"><div class="card-head">🚢 布袋船优化排计划（BSA利用率最大化 + 中转时间最短）</div>')
A('<div class="kpi-grid">')
A(f'<div class="kpi kpi-green"><div class="kpi-val">{util_pct:.0f}%</div><div class="kpi-lbl">BSA利用率</div></div>')
A(f'<div class="kpi"><div class="kpi-val">{avg_wait:.1f}天</div><div class="kpi-lbl">加权平均等待</div></div>')
A(f'<div class="kpi kpi-green"><div class="kpi-val">{total_assigned_teu}</div><div class="kpi-lbl">已安排TEU</div></div>')
unassigned = total_cargo_teu - total_assigned_teu
if unassigned > 0:
    A(f'<div class="kpi kpi-orange"><div class="kpi-val">{unassigned}</div><div class="kpi-lbl">未安排TEU</div></div>')
A('</div>')  # kpi-grid
A('<div class="legend-bar">')
A('🟦 <span class="color-dot" style="background:#fff"></span> 白底=干线船(国内→巴生卸货) ')
A('🟧 <span class="color-dot" style="background:#fff3e0"></span> 橙底=布袋船(巴生→红海装货·接驳) ')
A('🟩 <span class="color-dot" style="background:#fce4ec"></span> 粉底=直航(国内→红海不经巴生) ')
A('<span style="margin-left:auto;color:#888">算法: 加权贪心 | 时间权重×'+str(ALPHA)+' | 填充权重×'+str(BETA)+' | 最小操作窗口≥'+str(MIN_WAIT_DAYS)+'天</span>')
A('</div></div>')  # card

# Section 1: Ship inventory
A('<div class="card"><div class="card-head">📋 全部船期 R51~R77</div>')
A('<div style="overflow-x:auto"><table><thead><tr><th>R#</th><th>RGB</th><th>类型</th><th>VESSEL</th><th>SVC</th><th>TEU/BSA</th><th>PKG</th><th>SOK</th><th>JED</th><th>MUN</th><th>NGB</th><th>REMARKS</th></tr></thead><tbody>')
for s in ships:
    d=s['dst']; bg='w' if s['type']=='WHITE' else ('f' if s['type']=='FEEDER' else ('p' if s['type']=='PINK' else 'o'))
    lab={'WHITE':'白(干线)','FEEDER':'橙(布袋)','PINK':'粉(直航)'}.get(s['type'],'?')
    A(f"<tr><td>{s['row']}</td><td style='font-size:10px;monospace'>{s['rgb']}</td>")
    A(f"<td><span class='badge bg-{bg}'>{lab}</span></td><td><strong>{s['vessel']}</strong></td><td>{s['svc']}</td>")
    A(f"<td class='teu-n'>{s['teu']}</td><td class='date-f'>{s['pkg']}</td>")
    A(f"<td class='date-f'>{d.get('SOK','')}</td><td class='date-f'>{d.get('JED','')}</td><td class='date-f'>{d.get('MUN','')}</td>")
    A(f"<td class='date-f'>{d.get('NGB','')}</td><td class='rmks'>{s['rmk']}</td></tr>")
A('</tbody></table></div></div>')

# Section 2: Optimized assignment plan grouped by mainline
A('<div class="card"><div class="card-head">🔗 优化衔接计划（按干线船分组）</div>')
A('<div class="card-sub">按「评分= -时间惩罚 + 填充奖励」排序的贪心分配结果</div><div style="padding:12px">')

# Group by mainline
by_mainline = defaultdict(list)
for a in final_assignments:
    key = a['mainline']['vessel']+str(a['mainline']['row'])
    by_mainline[key].append(a)

# Sort mainlines by their first arrival date
ml_order = sorted(by_mainline.keys(), 
                  key=lambda k: parse_dt(by_mainline[k][0]['mainline']['pkg']) or datetime.max)

for mk in ml_order:
    assigns = by_mainline[mk]
    m = assigns[0]['mainline']
    rem = mainline_remaining.get(mk, 0)
    orig = m['teu']
    done = orig - rem
    
    best_wait = min(a['wait_days'] for a in assigns)
    wc = 'good' if best_wait<=3 else ('warn' if best_wait<=5 else 'bad')
    
    A('<div class="asgn-card">')
    A(f"<div class='asgn-head'>")
    A(f"<span style='background:#c8e6c9;padding:4px 10px;border-radius:4px;'>{m['vessel']}</span>")
    A(f"<small>ETA PKG: <code>{m['pkg']}</code> | 货量:{orig}TEU | 已安排:{done}TEU | 剩余:{rem}TEU</small>")
    A(f"<span style='margin-left:auto' class='{wc}'>最快等 {best_wait} 天</span>")
    A("</div>")  # head
    A("<div class='asgn-body'>")
    
    # Sort assigns by wait_days within group
    assigns.sort(key=lambda x: x['wait_days'])
    
    for a in assigns:
        f=a['feeder']; dd=f['dst']; wcls='good' if a['wait_days']<=3 else ('warn' if a['wait_days']<=5 else 'bad')
        
        # Feeder utilization after this assignment
        fk=f['vessel']+str(f['row']); cap=a['bsa_cap']; rem_f=feeder_remaining[fk]; used_f=cap-rem_f
        pct=(used_f/cap*100) if cap>0 else 0
        bar_cls='green' if pct>80 else ('yellow' if pct>40 else 'red')
        
        parts=[]; 
        for p in ['SOK','JED','MUN','NGB']: 
            if dd.get(p): parts.append(p+':'+dd[p])
        dest_s=', '.join(parts) if parts else '-'
        
        A("<div class='asgn-row'>")
        A(f"<span style='width:180px'><b>{f['vessel']}</b> ({f['svc']})</span>")
        A(f"<span class='date-f' style='width:80px'>ETD {f['pkg']}</span>")
        A(f"<span class='{wcls}' style='width:60px;text-align:center'>等{a['wait_days']}天</span>")
        A(f"<span style='width:70px;text-align:right'><b>+{a['assigned_teu']}TEU</b></span>")
        A(f"<span class='progress-wrap'><span class='progress-fill {bar_cls}' style='width:{pct:.0f}%'></span></span>")
        A(f"<span style='width:40px;text-align:right'>{pct:.0f}%</span>")
        A(f"<span style='color:#666;margin-left:6px'>→{dest_s}</span>")
        A("</div>")  # row
    A("</div>")  # body
    A("</div>\n")  # card

A('</div></div>')  # section

# Section 3: Global ranking table
A('<div class="card"><div class="card-head">📊 衔接计划表（按干线船 ETA PKG 时间排序）</div>')
A('<div class="card-sub">所有有效衔接按干线船到达巴生时间先后排列</div>')
A('<div style="overflow-x:auto"><table><thead>')
A('<tr><th>#</th><th>干线船(白)</th><th>ETA PKG</th><th>布袋船(橙)</th><th>ETD PKG</th><th>等天</th><th>安排TEU</th><th>状态</th><th>BSA容量</th><th>填充率</th><th>SVC</th></tr>')
A('</thead><tbody>')

ranked = sorted(final_assignments, key=lambda x: parse_dt(x['mainline']['pkg']) or datetime.max)
for i,a in enumerate(ranked,1):
    m=a['mainline']; f=a['feeder']; dd=f['dst']
    wcls='good' if a['wait_days']<=3 else ('warn' if a['wait_days']<=5 else 'bad')
    status='✅快转' if wcls=='good' else ('⚠️正常' if wcls=='warn' else '🔴较慢')
    fk=f['vessel']+str(f['row']); cap=a['bsa_cap']; rem_f=feeder_remaining[fk]; used_f=cap-rem_f
    fill_pct=(used_f/cap*100) if cap>0 else 0
    sc_tag='good' if a['score']>50 else ('warn' if a['score']>0 else 'bad')
    
    parts=[]
    for p in ['SOK','JED','MUN','NGB']:
        if dd.get(p): parts.append(p+':'+dd[p])
    dest_str=', '.join(parts) if parts else '-'
    
    A(f"<tr><td>{i}</td><td><b>{m['vessel']}</b></td><td class='date-f'>{m['pkg']}</td>")
    A(f"<td><b>{f['vessel']}</b></td><td class='date-f'>{f['pkg']}</td>")
    A(f"<td class='{wcls}'>{a['wait_days']}</td><td class='teu-n'>{a['assigned_teu']}</td>")
    A(f"<td>{status}</td><td class='teu-n'>{cap}</td><td>{fill_pct:.0f}%</td>")
    A(f"<td>{f['svc']} → {dest_str}</td></tr>")

A('</tbody></table></div></div>')

# Section 4: BSA utilization per feeder
A('<div class="card"><div class="card-head">📦 各布袋船 BSA 利用情况</div>')
A('<div style="overflow-x:auto"><table><thead><tr><th>#</th><th>布袋船(橙)</th><th>SVC</th><th>BSA容量</th><th>已安排</th><th>剩余</th><th>利用率</th><th>进度</th><th>PKGETD</th><th>目的港</th></tr></thead><tbody>')

feeders_sorted = sorted(feeders, key=lambda f: f['pkg_dt'] or datetime.max)
for i,f in enumerate(feeders_sorted,1):
    fk=f['vessel']+str(f['row']); cap=f['teu']
    rem=feeder_remaining.get(fk,cap); used=cap-rem
    pct=(used/cap*100) if cap>0 else 0
    bar_cls='green' if pct>75 else ('yellow' if pct>30 else 'red')
    dd=f['dst']
    parts=[]
    for p in ['SOK','JED','MUN','NGB']: 
        if dd.get(p): parts.append(p+':'+dd[p])
    dest_s=', '.join(parts) if parts else '-'
    
    A(f"<tr><td>{i}</td><td><b>{f['vessel']}</b></td><td>{f['svc']}</td>")
    A(f"<td class='teu-n'>{cap}</td><td class='teu-n'>{used}</td><td class='teu-n'>{rem}</td>")
    A(f"<td>{pct:.0f}%</td>")
    A(f"<td><div class='progress-wrap'><span class='progress-fill {bar_cls}' style='width:{min(pct,100):.0f}%'></span></div></td>")
    A(f"<td class='date-f'>{f['pkg']}</td><td>{dest_s}</td></tr>")

A('</tbody></table></div></div>')

A('</body></html>')

H='\n'.join(L)
out=r'c:\Users\leahliu\WorkBuddy\20260525175100\budai_plan.html'
with open(out,'w',encoding='utf-8') as f: f.write(H)
print(f"\nOutput: {out}")
