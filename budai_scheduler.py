"""
布袋船排计划工具 v2.0
======================
逻辑：
- 白底(国内→巴生) 到达后，货需要橙底(布袋船/巴生→红海) 接走
- 每条白底船的货量 = EFF.TEUS（假设满载）
- 每条橙底船的BSA舱位 = EFF.TEUS（可用容量）
- 目标：等待时间最短 + 舱位匹配最优
- 货物可拆分给多条布袋船
"""

from openpyxl import load_workbook
from datetime import datetime, timedelta

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']
header_row = 7

# Build column map
col_map={}
for c in range(1,sheet.max_column+1):
    h=sheet.cell(row=header_row,column=c).value
    if h: col_map[h.strip().upper()]=c

pkg_col=col_map.get('PKG',21)
teu_col=col_map.get('EFF. TEUS',12)
vessel_col=col_map.get('CUL VESSELS',3)
week_col=col_map.get('WEEK',2)
service_col=col_map.get('SERVICES',5)
voyage_col=col_map.get('VOYAGE .No',4)
remarks_col=col_map.get('REMARKS',34)

dest_cols={}
for pn in ['SOK','JED','MUN','NGB']:
    if pn in col_map: dest_cols[pn]=col_map[pn]

vol_wb_col=col_map.get('DISCHARGE VOL')  # WB PKG CARGO VOLUME

def fmt(d):
    if d is None: return ''
    if isinstance(d,datetime): return d.strftime('%Y-%m-%d')
    s=str(d)[:10]; return s if s else ''

def get_rgb(cell):
    try: r=str(cell.fill.fgColor.rgb); return r.upper() if len(r)==8 else '00000000'
    except: return '00000000'

def classify(rgb):
    if rgb=='00000000': return 'WHITE'
    if 'DAF2D0' in rgb: return 'PINK'  # 直航 国内→红海
    return 'FEEDER'                     # 橙色 布袋船 巴生→红海

def parse_dt(s):
    if not s: return None
    try: return datetime.strptime(str(s)[:10],'%Y-%m-%d')
    except: return None

def safe_int(v,default=0):
    try: return int(float(v))
    except: return default

# Parse R51-R77
all_rows=[]
for row in range(51,78):
    vc=sheet.cell(row=row,column=vessel_col); v=str(vc.value or '').strip()
    if not v: continue
    
    rgb=get_rgb(vc); ctype=classify(rgb)
    
    pkg=sheet.cell(row=row,column=pkg_col).value
    teu=sheet.cell(row=row,column=teu_col).value
    vol_wb=sheet.cell(row=row,column=vol_wb_col).value if vol_wb_col else None
    wk=sheet.cell(row=row,column=week_col).value; svc=sheet.cell(row=row,column=service_col).value
    voy=sheet.cell(row=row,column=voyage_col).value
    rmk=sheet.cell(row=row,column=remarks_col).value
    
    dst={}
    for pn,pc in dest_cols.items():
        val=sheet.cell(row=row,column=pc).value
        if val: dst[pn]=fmt(val)
    
    all_rows.append({
        'row':row,'vessel':v,'wk':str(wk or ''),'svc':str(svc or ''),
        'voy':str(voy or ''),'pkg':fmt(pkg),'teu':safe_int(teu),
        'vol_wb':safe_int(vol_wb) if vol_wb else None,
        'rgb':rgb,'ctype':ctype,
        'rmk':(str(rmk)[:80] if rmk else ''),'dst':dst
    })

white=[r for r in all_rows if r['ctype']=='WHITE']
feeders=[r for r in all_rows if r['ctype']=='FEEDER']  # orange = 布袋船
pink=[r for r in all_rows if r['ctype']=='PINK']         # 直航，不参与衔接

print("White(mainline): %d, Orange(feeder): %d, Pink(direct): %d" % (len(white),len(feeders),len(pink)))
for w in white: print("  W R%02d: %-24s PKG=%s TEU=%d vol=%s" % (w['row'],w['vessel'],w['pkg'],w['teu'],w['vol_wb']))
for f in feeders: print("  F R%02d: %-24s PKG=%s TEU=%d" % (f['row'],f['vessel'],f['pkg'],f['teu']))
for p in pink:   print("  P R%02d: %-24s PKG=%s TEU=%d" % (p['row'],p['vessel'],p['pkg'],p['teu']))

# ============================================================
# SCHEDULING ALGORITHM
# ============================================================
# For each white ship arriving at PKG:
#   cargo = vol_wb if available, else teu (assume full)
#   Find available feeders departing >= min_window days after arrival
#   Assign cargo to feeders by priority: shortest wait first
#   Track remaining feeder capacity

MIN_WINDOW = 1  # minimum operation days at PKG

# Initialize feeder remaining capacity
feeder_cap = {f['row']: f['teu'] for f in feeders}

# Build schedule
schedule = []  # list of assignment records
assignment_id = 1

for w in sorted(white, key=lambda x: x['pkg']):
    eta = parse_dt(w['pkg'])
    if not eta:
        schedule.append({
            'id': assignment_id, 'white': w, 'feeder': None,
            'wait': None, 'cargo': 0, 'assigned_teu': 0,
            'status': 'NO_PKG_DATE'
        })
        assignment_id += 1
        continue
    
    # Determine cargo volume
    cargo = w['vol_wb'] if w['vol_wb'] and w['vol_wb'] > 0 else w['teu']
    remaining_cargo = cargo
    
    # Find eligible feeders sorted by wait time
    candidates = []
    for f in feeders:
        etd = parse_dt(f['pkg'])
        if not etd: continue
        delta = (etd - eta).days
        if delta < MIN_WINDOW: continue
        avail = feeder_cap[f['row']]
        if avail <= 0: continue
        candidates.append((delta, f, avail))
    
    candidates.sort(key=lambda x: x[0])  # shortest wait first
    
    # Greedy assign
    assignments_for_this_ship = []
    for delta, f, avail in candidates:
        if remaining_cargo <= 0: break
        assign_teu = min(remaining_cargo, avail)
        
        status = 'good' if delta <= 3 else ('warn' if delta <= 5 else 'slow')
        
        rec = {
            'id': assignment_id, 'white': w, 'feeder': f,
            'wait': delta, 'cargo_total': cargo,
            'assign_teu': assign_teu, 'remaining_after': max(0, remaining_cargo - assign_teu),
            'status': status, 'feeder_avail_before': avail,
            'feeder_avail_after': avail - assign_teu
        }
        assignments_for_this_ship.append(rec)
        schedule.append(rec)
        assignment_id += 1
        
        # Update state
        remaining_cargo -= assign_teu
        feeder_cap[f['row']] -= assign_teu
    
    if not assignments_for_this_ship:
        schedule.append({
            'id': assignment_id, 'white': w, 'feeder': None,
            'wait': None, 'cargo_total': cargo, 'assign_teu': 0,
            'remaining_after': cargo, 'status': 'NO_FEEDER'
        })
        assignment_id += 1
    elif remaining_cargo > 0:
        schedule.append({
            'id': assignment_id, 'white': w, 'feeder': None,
            'wait': None, 'cargo_total': cargo, 'assign_teu': 0,
            'remaining_after': remaining_cargo, 'status': 'PARTIAL_UNASSIGNED',
            'note': '%d TEU未安排' % remaining_cargo
        })
        assignment_id += 1

# ============================================================
# BUILD HTML REPORT
# ============================================================
lines=[]; A=lines.append

A('<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">')
A('<title>🚢 布袋船排计划 - R51~R77</title>')
A('''<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:#f5f6fa;padding:20px;color:#333}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 10px rgba(0,0,0,.08);margin-bottom:20px;overflow:hidden}
.hd{padding:14px 20px;font-weight:bold;font-size:16px;border-bottom:1px solid #eee;display:flex;align-items:center;gap:10px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#fafbfc;padding:9px 11px;text-align:left;font-weight:600;border-bottom:2px solid #dde1e6;position:sticky;top:0;white-space:nowrap}
td{padding:7px 11px;border-bottom:1px solid #f0f0f0;white-space:nowrap}
tr:hover td{background:#f8f9ff}
.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:12px}
.bg-w{background:#e8eaed;color:#444} .bg-p{background:#fce4ec;color:#ad1457} .bg-f{background:#fff3e0;color:#e65100}
.date-f{font-family:"Courier New",monospace;font-size:12px}
.teu-n{font-weight:600;text-align:right;font-family:Consolas,"Courier New",monospace}
.good{color:#1565c0;font-weight:600} .warn{color:#e65100;font-weight:600} .slow{color:#c62828;font-weight:600}
.rmks{color:#777;font-style:italic;font-size:11px;max-width:220px;overflow:hidden;text-overflow:ellipsis}
.summary{display:flex;gap:28px;padding:18px 20px;flex-wrap:wrap}
.s-item .n{font-size:30px;font-weight:bold;color:#1565c0}
.s-item .l{font-size:12px;color:#666;margin-top:2px}
.plan-card{border:1px solid #e0e4ea;border-radius:10px;margin-bottom:14px;overflow:hidden}
.plan-head{padding:12px 16px;display:flex;align-items:center;gap:14px;font-weight:bold;font-size:15px}
.plan-body{padding:12px 16px}
.assign-row{display:flex;align-items:center;padding:7px 10px;border-bottom:1px dashed #e8e8e8;border-radius:5px;margin:3px 0;font-size:12.5px;transition:background .15s}
.assign-row:hover{background:#f0f7ff}
.assign-row:last-child{border-bottom:none}
.unassigned{background:#fff8e1;border-color:#ffecb3}
.legend-bar{padding:12px 20px;display:flex;gap:22px;font-size:12px;border-top:1px solid #eee;background:#fafafa;flex-wrap:wrap}
.color-sq{width:18px;height:18px;border-radius:4px;border:1px solid #ccc;display:inline-block;vertical-align:middle}
.stat-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;padding:16px 20px}
.stat-box{background:#f8f9fc;border:1px solid #e8ecf0;border-radius:8px;padding:12px;text-align:center}
.stat-box .val{font-size:26px;font-weight:bold}
.stat-box .lbl{font-size:11px;color:#666;margin-top:4px}
.arrow{color:#999;font-size:16px;margin:0 6px}
</style></head><body>''')

# === HEADER CARD ===
A('<div class="card">')
A('<div class="hd">🚢 布袋船排计划 — R51~R77</div>')

total_cargo = sum(s.get('cargo_total',0) for s in schedule)
assigned_teu = sum(s.get('assign_teu',0) for s in schedule)
unassigned_teu = sum(s.get('remaining_after',0) for s in schedule if s.get('status')=='PARTIAL_UNASSIGNED' or s.get('status')=='NO_FEEDER')

A('<div class="summary">')
A('<div class="s-item"><div class="n">' + str(len(white)) + '</div><div class="l">白底干线船<br>(国内→巴生)</div></div>')
A('<div class="s-item"><div class="n">' + str(len(feeders)) + '</div><div class="l">橙底布袋船<br>(巴生→红海)</div></div>')
A('<div class="s-item"><div class="n">' + str(total_cargo) + '</div><div class="l">总货量(TEU)<br>(假设满载)</div></div>')
A('<div class="s-item"><div class="n">' + str(assigned_teu) + '</div><div class="l">已安排TEU</div></div>')
if unassigned_teu > 0:
    A('<div class="s-item"><div class="n" style="color:#c62828">' + str(unassigned_teu) + '</div><div class="l">未安排TEU</div></div>')
A('</div>')

A('<div class="legend-bar">')
A('<span><span class="color-sq" style="background:#fff"></span> 白底= 干线船（国内→巴生卸货）</span>')
A('<span><span class="color-sq" style="background:#DAF2D0"></span> 粉底= 直航（国内→红海，不参与）</span>')
A('<span><span class="color-sq" style="background:#FBE2D5"></span> 橙底= 布袋船（巴生装货→红海）✅</span>')
A('<span style="margin-left:auto;color:#555">操作窗口 ≥ 1天 | 快转 ≤3天 🔵 | 正常 4-5天 🟡 | >5天 🔴</span>')
A('</div></div>')

# === SHIP LIST ===
A('<div class="card"><div class="hd">📋 全部船期（R51~R77 颜色分类确认）</div>')
A('<div style="overflow-x:auto"><table><thead><tr><th>R#</th><th>RGB</th><th>类型</th><th>VESSEL</th><th>SVC</th><th>TEU</th><th>VOL_WB</th><th>PKG</th><th>SOK</th><th>JED</th><th>MUN</th><th>NGB</th><th>REMARKS</th></tr></thead><tbody>')
for r in all_rows:
    bg='w' if r['ctype']=='WHITE' else ('f' if r['ctype']=='FEEDER' else 'p')
    lab='白底' if bg=='w' else ('布袋' if bg=='f' else '直航')
    d=r['dst']
    A("<tr><td>%d</td><td style='font-size:10px;monospace'>%s</td>" % (r['row'],r['rgb']))
    A("<td><span class='badge bg-%s'>%s</span></td>" % (bg,lab))
    A("<td><strong>%s</strong></td><td>%s</td><td class='teu-n'>%s</td>" % (r['vessel'],str(r['svc']),str(r['teu']) if r['teu'] else '-'))
    A("<td class='teu-n'>%s</td><td class='date-f'>%s</td>" % (str(r['vol_wb']) if r['vol_wb'] else '-',r['pkg']))
    A("<td class='date-f'>%s</td><td class='date-f'>%s</td><td class='date-f'>%s</td><td class='date-f'>%s</td>" % (d.get('SOK',''),d.get('JED',''),d.get('MUN',''),d.get('NGB','')))
    A("<td class='rmks'>%s</td></tr>" % r['rmk'])
A('</tbody></table></div></div>')

# === SCHEDULE PLAN ===
A("<div class='card'><div class='hd'>📌 布袋船衔接计划（按干线船分组）</div>")
A('<div style="padding:16px">')

# Group schedule by white ship
from collections import defaultdict
by_white = defaultdict(list)
for s in schedule:
    key = s['white']['row']
    by_white[key].append(s)

for w in white:
    assigns = by_white.get(w['row'],[])
    
    A("<div class='plan-card'>")
    
    # Header
    total_assign = sum(a.get('assign_teu',0) for a in assigns)
    unassign = assigns[-1].get('remaining_after',0) if assigns else 0
    head_status = ''
    if not any(a.get('assign_teu',0)>0 for a in assigns):
        head_status = '<span style="color:#c62828;margin-left:auto">❌ 无可接驳船</span>'
    elif unassign > 0:
        head_status = '<span style="color:#e65100;margin-left:auto">⚠️ 剩余%d TEU 未安排</span>' % unassign
    else:
        head_status = '<span style="color:#2e7d32;margin-left:auto">✅ 全部已安排</span>'
    
    A("<div class='plan-head'><span style='background:#e8f5e9;padding:4px 10px;border-radius:4px'>%s</span>" % w['vessel'])
    A("<small>ETA PKG: <code>%s</code></small>" % w['pkg'])
    A("<small>| %s | 货量:<strong>%d TEU</strong></small>" % (w['svc'], w['vol_wb'] if w['vol_wb'] and w['vol_wb']>0 else w['teu']))
    A(head_status)
    A("</div>")  # plan-head
    
    A("<div class='plan-body'>")
    if assigns:
        for a in assigns:
            if not a.get('feeder'):
                # Unassigned entry
                st = a.get('status','')
                note = a.get('note','')
                if st == 'NO_FEEDER':
                    A("<div class='assign-row unassigned'><span style='color:#c62828'>⚠️ 无满足≥1天操作窗口的布袋船可接，%d TEU 待安排</span></div>" % a.get('cargo_total',0))
                elif st == 'PARTIAL_UNASSIGNED':
                    A("<div class='assign-row unassigned'><span style='color:#e65100'>⚠️ %s</span></div>" % note)
                elif st == 'NO_PKG_DATE':
                    A("<div class='assign-row unassigned'><span style='color:#999'>ℹ️ 该船无 PKG ETA 日期</span></div>")
                continue
            
            f = a['feeder']
            fdst=f['dst']
            dest_parts=[]
            for p in ['SOK','JED','MUN','NGB']:
                if fdst.get(p): dest_parts.append(p+':'+fdst[p])
            dest_str=', '.join(dest_parts) if dest_parts else '-'
            
            cls = a['status']
            wait_label = '%d天 ✅快转' if cls=='good' else ('%d天 ⚠️正常' if cls=='warn' else '%d天 🔴较慢')
            
            A("<div class='assign-row'>")
            A("<span style='width:200px'><strong>%s</strong> <small>(%s)</small></span>" % (f['vessel'], f['svc']))
            A("<span class='date-f' style='width:90px'>ETD PKG %s</span>" % f['pkg'])
            A("<span class='%s' style='width:90px;text-align:center'>等%s</span>" % (cls, wait_label % a['wait']))
            A("<span class='arrow'>→</span>")
            A("<span class='teu-n' style='width:70px'>接 %d TEU</span>" % a['assign_teu'])
            A("<span style='color:#888;width:50px'>/</span>")
            A("<span class='teu-n' style='width:70px'>余 %d</span>" % a['feeder_avail_after'])
            A("<span style='color:#666;margin-left:8px;font-size:11.5px'>→ %s</span>" % dest_str)
            A("</div>")
    A("</div></div>\n")  # plan-body, plan-card

A("</div></div>")

# === SUMMARY TABLE (flat sorted by wait time) ===
A("<div class='card'><div class='hd'>📊 衔接汇总表（全局排序：等待时间最短优先）</div>")
A("<div style='overflow-x:auto'><table><thead><tr><th>#</th><th>干线船(白)</th><th>ETA PKG</th><th>货量TEU</th><th>布袋船(橙)</th><th>ETD PKG</th><th>等待</th><th>状态</th><th>安排TEU</th><th>布袋余位</th><th>目的港</th></tr></thead><tbody>")

flat=[]
for s in schedule:
    if s.get('feeder') and s.get('assign_teu',0)>0:
        flat.append(s)
flat.sort(key=lambda x:(x['wait'],x['white']['pkg']))

for i,s in enumerate(flat,1):
    w=s['white']; f=s['feeder']; fd=f['dst']
    dp=[]
    for p in ['SOK','JED','MUN','NGB']:
        if fd.get(p): dp.append(p+':'+fd[p])
    ds=', '.join(dp) if dp else '-'
    st='✅快转' if s['status']=='good' else ('⚠️正常' if s['status']=='warn' else '🔴较慢')
    A("<tr><td>%d</td><td><strong>%s</strong></td><td class='date-f'>%s</td><td class='teu-n'>%d</td>" % (i,w['vessel'],w['pkg'],s['cargo_total']))
    A("<td><strong>%s</strong></td><td class='date-f'>%s</td><td class='%s'>%d天</td>" % (f['vessel'],f['pkg'],s['status'],s['wait']))
    A("<td>%s</td><td class='teu-n'>%d</td><td class='teu-n'>%d</td><td>%s</td></tr>" % (st,s['assign_teu'],s['feeder_avail_after'],ds))

A("</tbody></table></div></div>")

# Feeder utilization summary
A("<div class='card'><div class='hd'>📦 布袋船舱位使用情况</div>")
A("<div style='overflow-x:auto'><table><thead><tr><th>#</th><th>布袋船(橙)</th><th>SVC</th><th>总BSA(TEU)</th><th>已用</th><th>剩余</th><th>利用率</th><th>ETD PKG</th><th>去向</th></tr></thead><tbody>")
for fi,f in enumerate(sorted(feeders,key=lambda x:x['pkg']),1):
    used = f['teu'] - feeder_cap[f['row']]
    remain = feeder_cap[f['row']]
    pct = int(used/f['teu']*100) if f['teu']>0 else 0
    bar_color='#4caf50' if pct<70 else ('#ff9800' if pct<95 else '#f44336')
    fd=f['dst']; dp=[p+':'+fd[p] for p in ['SOK','JED','MUN','NGB'] if fd.get(p)]
    
    A("<tr><td>%d</td><td><strong>%s</strong></td><td>%s</td><td class='teu-n'>%d</td>" % (fi,f['vessel'],f['svc'],f['teu']))
    A("<td class='teu-n'>%d</td><td class='teu-n'>%d</td>" % (used,remain))
    A("<td><div style='width:80px;background:#eee;border-radius:4px;height:16px;overflow:hidden'><div style='height:100%%;width:%d%%;background:%s;border-radius:4px'></div></div><small>%d%%</small></td>" % (pct,bar_color,pct))
    A("<td class='date-f'>%s</td><td>%s</td></td></tr>" % (f['pkg'],', '.join(dp) if dp else '-'))

A("</tbody></table></div></div>")

A("</body></html>")

H='\n'.join(lines)
out=r'c:\Users\leahliu\WorkBuddy\20260525175100\budai_plan.html'
with open(out,'w',encoding='utf-8') as fp:
    fp.write(H)

print("\nDone! Output:", out)
print("Total cargo: %d, Assigned: %d, Unassigned: %d" % (total_cargo, assigned_teu, unassigned_teu))
