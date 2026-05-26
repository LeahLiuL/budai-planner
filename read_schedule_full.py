from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

# Key column indices (1-indexed)
pkg_col = col_map.get('PKG', 21)
pkge_col = col_map.get('PKG EB', 29)
teu_col = col_map.get('EFF. TEUS', 12)
vol_col_wb = col_map.get('DISCHARGE VOL', 30)   # WB PKG CARGO VOL
vol_col_eb = col_map.get('LOAD VOL', 31)        # EB PKG LOAD VOL
vessel_col = col_map.get('CUL VESSELS', 3)
week_col = col_map.get('WEEK', 2)
service_col = col_map.get('SERVICES', 5)
voyage_col = col_map.get('VOYAGE .NO', 4)

# Also get destination port columns
dest_ports = {}
for port_name, col_key in [('SOK', 'SOK'), ('JED', 'JED'), ('MUN', 'MUN'), ('NGB', 'NGB')]:
    c = col_map.get(col_key, None)
    if c:
        dest_ports[port_name] = c

# Parse all data rows
white_rows = []   # main line (domestic -> PKG) - white/no fill
pink_rows = []    # feeder (PKG -> dest) - pink

for row in range(header_row + 1, sheet.max_row + 1):
    vessel_cell = sheet.cell(row=row, column=vessel_col)
    vessel = str(vessel_cell.value or '').strip()
    
    if not vessel:
        continue
    
    fill = vessel_cell.fill
    rgb_raw = str(fill.fgColor.rgb) if fill and fill.fgColor and fill.fgColor.rgb else '00000000'
    
    # Normalize RGB - handle different formats
    rgb = rgb_raw.replace('00', '', 1) if len(rgb_raw) == 8 else rgb_raw
    
    # Color classification
    pink_colors = ['DAF2D0', 'FBE2D5', 'CAEDFB', 'FFBF', 'BFBFBF']  # various pink/orange/gray tones
    is_pink = any(p in rgb.upper() for p in pink_colors) or rgb_raw.startswith('FFDA') or rgb_raw.startswith('FFF')
    
    week = sheet.cell(row=row, column=week_col).value
    svc = sheet.cell(row=row, column=service_col).value
    voyage = sheet.cell(row=row, column=voyage_col).value
    pkg_eta = sheet.cell(row=row, column=pkg_col).value if pkg_col else None
    pkge_eta = sheet.cell(row=row, column=pkge_col).value if pkge_col else None
    teu = sheet.cell(row=row, column=teu_col).value if teu_col else None
    vol_wb = sheet.cell(row=row, column=vol_col_wb).value if vol_col_wb else None
    vol_eb = sheet.cell(row=row, column=vol_col_eb).value if vol_col_eb else None
    
    remarks = sheet.cell(row=row, column=col_map.get('REMARKS', 34)).value if 'REMARKS' in col_map else None
    
    # Destination ports ETA
    dest_etas = {}
    for pname, pc in dest_ports.items():
        val = sheet.cell(row=row, column=pc).value
        if val:
            dest_etas[pname] = val
    
    def fmt_date(d):
        if d is None:
            return ''
        if isinstance(d, datetime):
            return d.strftime('%Y-%m-%d')
        return str(d)[:10]
    
    row_data = {
        'row': row,
        'vessel': vessel,
        'week': str(week) if week else '',
        'service': str(svc) if svc else '',
        'voyage': str(voyage) if voyage else '',
        'eta_pkg': fmt_date(pkg_eta),
        'eta_pkge': fmt_date(pkge_eta),
        'teu': int(teu) if teu and str(teu).isdigit() else teu,
        'vol_wb': vol_wb,
        'vol_eb': vol_eb,
        'color_rgb': rgb_raw,
        'remarks': str(remarks)[:80] if remarks else '',
        'dest_etas': {k: fmt_date(v) for k, v in dest_etas.items()}
    }
    
    if is_pink:
        pink_rows.append(row_data)
    else:
        white_rows.append(row_data)

# Build HTML report
html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>REX & Feeder Schedule 数据提取报告 - 25_May</title>
<style>
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family: -apple-system, "Segoe UI", "Microsoft YaHei", sans-serif; background:#f0f2f5; padding:20px; color:#333; }
h1 { font-size:20px; margin-bottom:8px; color:#1a73e8; }
.subtitle { color:#666; margin-bottom:16px; font-size:13px; }
.card { background:#fff; border-radius:8px; box-shadow:0 1px 4px rgba(0,0,0,.1); margin-bottom:16px; overflow:hidden; }
.card-header { padding:12px 16px; font-weight:bold; font-size:15px; display:flex; align-items:center; justify-content:space-between; }
.badge { display:inline-block; padding:2px 10px; border-radius:12px; font-size:12px; font-weight:normal; }
.badge-white { background:#e8eaed; color:#333; }
.badge-pink { background:#fce4ec; color:#c2185b; }
.card-body { padding:0; }
table { width:100%; border-collapse:collapse; font-size:12px; }
th { background:#f8f9fa; padding:6px 8px; text-align:left; font-weight:600; border-bottom:2px solid #dee2e6; position:sticky; top:0; white-space:nowrap; }
td { padding:5px 8px; border-bottom:1px solid #eee; white-space:nowrap; max-width:200px; overflow:hidden; text-overflow:ellipsis; }
tr:hover td { background:#f0f7ff; }
.date-cell { font-family:Consolas,"Courier New",monospace; font-size:11px; }
.teu-cell { font-weight:600; text-align:right; }
.wait-ok { color:#2e7d32; font-weight:600; }
.wait-warn { color:#e65100; font-weight:600; }
.wait-bad { color:#c62828; font-weight:600; }
.remark-cell { max-width:300px; color:#666; font-style:italic; font-size:11px; }
.summary-grid { display:flex; gap:12px; padding:16px; flex-wrap:wrap; }
.summary-item { min-width:140px; }
.summary-num { font-size:24px; font-weight:bold; color:#1a73e8; }
.summary-label { font-size:11px; color:#666; }
.legend { padding:12px 16px; display:flex; gap:20px; font-size:12px; border-top:1px solid #eee; }
.legend-item { display:flex; align-items:center; gap:6px; }
.legend-color { width:18px; height:18px; border-radius:3px; border:1px solid #ddd; }
.legend-white { background:#fff; }
.legend-pink { background:#DAF2D0; }
.match-section { padding:16px; }
.match-card { border:1px solid #ddd; border-radius:6px; margin-bottom:12px; overflow:hidden; }
.match-header { background:#f0f7ff; padding:8px 12px; font-weight:bold; font-size:13px; display:flex; gap:12px; align-items:center; }
.match-body { padding:8px 12px; }
.match-row { display:flex; align-items:center; padding:4px 0; gap:8px; font-size:12px; border-bottom:1px dotted #eee; }
.match-row:last-child { border-bottom:none; }
</style>
</head>
<body>

<div class="card">
<div class="card-header">
  <span>📋 REX & FEEDER SCHEDULE — 数据提取报告</span>
  <span style="font-size:12px;font-weight:normal;color:#666">Sheet: 25_May | 提取时间: 2026-05-25</span>
</div>
<div class="summary-grid">
  <div class="summary-item"><div class="summary-num">''' + str(len(white_rows)) + '''</div><div class="summary-label">干线船（白底）国内→PKG</div></div>
  <div class="summary-item"><div class="summary-num">''' + str(len(pink_rows)) + '''</div><div class="summary-label">接驳/布袋船（粉底）PKG→目的港</div></div>
  <div class="summary-item"><div class="summary-num">''' + str(len(white_rows)+len(pink_rows)) + '''</div><div class="summary-label">总计船次</div></div>
</div>
<div class="legend">
  <div class="legend-item"><div class="legend-color legend-white"></div> 白色底 = 干线船（国内港口→巴生PKG，卸货）</div>
  <div class="legend-item"><div class="legend-color legend-pink"></div> 粉色底 = 接驳船/布袋船（PKG→目的港，装货）</div>
</div>
</div>

'''

# Section 1: Mainline vessels (White)
html += '<div class="card"><div class="card-header"><span>🚢 干线船列表（白底 / 国内 → PKG）</span><span class="badge badge-white">' + str(len(white_rows)) + ' 艘</span></div>'
html += '<div style="overflow-x:auto;"><table><thead><tr><th>R#</th><th>WEEK</th><th>VESSEL</th><th>VOYAGE</th><th>SVC</th><th>TEU</th>'
html += '<th>ETA_PKG<br>国内→巴生</th><th>ETA_PKGE<br>红海→巴生</th>'
html += '<th>SOK</th><th>JED</th><th>MUN</th><th>NGB</th><th>REMARKS</th></tr></thead><tbody>\n'

for r in white_rows:
    dests = r['dest_etas']
    html += f"<tr><td>{r['row']}</td><td>{r['week']}</td><td><strong>{r['vessel']}</strong></td><td>{r['voyage']}</td><td>{r['service']}</td><td class='teu-cell'>{r['teu']}</td>"
    html += f"<td class='date-cell'>{r['eta_pkg']}</td><td class='date-cell'>{r['eta_pkge']}</td>"
    html += f"<td class='date-cell'>{dests.get('SOK','')}</td><td class='date-cell'>{dests.get('JED','')}</td><td class='date-cell'>{dests.get('MUN','')}</td><td class='date-cell'>{dests.get('NGB','')}</td>"
    html += f"<td class='remark-cell'>{r['remarks']}</td></tr>\n"

html += '</tbody></table></div></div>\n'

# Section 2: Feeder vessels (Pink)
html += '<div class="card"><div class="card-header"><span>⛴️ 接驳/布袋船列表（粉底 / PKG → 目的港）</span><span class="badge badge-pink">' + str(len(pink_rows)) + ' 艘</span></div>'
html += '<div style="overflow-x:auto;"><table><thead><tr><th>R#</th><th>WEEK</th><th>VESSEL</th><th>VOYAGE</th><th>SVC</th><th>TEU</th>'
html += '<th>ETA_PKG</th><th>ETA_PKGE</th>'
html += '<th>SOK</th><th>JED</th><th>MUN</th><th>NGB</th><th>COLOR_RGB</th><th>REMARKS</th></tr></thead><tbody>\n'

for r in pink_rows:
    dests = r['dest_etas']
    html += f"<tr><td>{r['row']}</td><td>{r['week']}</td><td><strong>{r['vessel']}</strong></td><td>{r['voyage']}</td><td>{r['service']}</td><td class='teu-cell'>{r['teu']}</td>"
    html += f"<td class='date-cell'>{r['eta_pkg']}</td><td class='date-cell'>{r['eta_pkge']}</td>"
    html += f"<td class='date-cell'>{dests.get('SOK','')}</td><td class='date-cell'>{dests.get('JED','')}</td><td class='date-cell'>{dests.get('MUN','')}</td><td class='date-cell'>{dests.get('NGB','')}</td>"
    html += f"<td style='font-size:10px;font-family:monospace;'>{r['color_rgb']}</td><td class='remark-cell'>{r['remarks']}</td></tr>\n"

html += '</tbody></table></div></div>\n'

# Section 3: Connection matching analysis
html += '<div class="card"><div class="card-header"><span>🔗 衔接匹配分析（干线船到PKG → 接驳船离开）</span></div><div class="match-section">'
html += '<p style="font-size:13px;color:#666;margin-bottom:12px;">规则：干线船 ETA_PKG 后 ≥ 1天操作窗口才能接上接驳船。按等待时间排序，最短优先。</p>\n'

def parse_dt(s):
    if not s or s == '' or s == 'X':
        return None
    try:
        return datetime.strptime(str(s)[:10], '%Y-%m-%d')
    except:
        try:
            return datetime.strptime(str(s), '%Y-%m-%d %H:%M:%S')
        except:
            return None

connection_count = 0
for mw in white_rows:
    mw_eta = parse_dt(mw['eta_pkg'])
    if not mw_eta:
        continue
    
    matches = []
    for mp in pink_rows:
        mp_etd = parse_dt(mp['eta_pkg'])
        if not mp_etd:
            continue
        
        delta_days = (mp_etd - mw_eta).days
        if delta_days >= 1:  # minimum operation window
            wait_class = 'wait-ok' if delta_days <= 3 else ('wait-warn' if delta_days <= 5 else 'wait-bad')
            matches.append({'feeder': mp, 'delta': delta_days, 'class': wait_class})
    
    if matches:
        matches.sort(key=lambda x: x['delta'])
        best = matches[0]
        connection_count += 1
        
        html += f"<div class='match-card'>"
        html += f"<div class='match-header'><span style='background:#e3f2fd;padding:3px 8px;border-radius:4px;'>{mw['vessel']}</span> <small>ETA PKG: <code>{mw['eta_pkg']}</code> | TEU:{mw['teu']} | {mw['service']}</small>"
        html += f" <span style='margin-left:auto;' class='{best['class']}'>最佳衔接等待 {best['delta']} 天</span></div>"
        html += "<div class='match-body'>"
        
        shown = 0
        for m in matches[:8]:  # top 8 matches per mainline ship
            f = m['feeder']
            html += f"<div class='match-row'><span style='width:180px;display:inline-block;'><b>{f['vessel']}</b> ({f['service']})</span>"
            html += f"<span class='date-cell' style='width:90px;'>ETD {f['eta_pkg']}</span>"
            html += f"<span class='{m['class']}' style='width:70px;text-align:center;'>等 {m['delta']}天</span>"
            html += f"<span style='width:60px;text-align:right;'>TEU {f['teu']}</span></div>"
            shown += 1
        
        if len(matches) > shown:
            html += f"<div style='font-size:11px;color:#999;padding:4px 0;'>... 还有 {len(matches)-shown} 条可衔接</div>"
        
        html += "</div></div>\n"

if connection_count == 0:
    html += '<p style="color:#c62828;">⚠️ 未找到有效衔接（可能日期格式需要调整或数据不完整）</p>\n'

html += '</div></div>\n'

html += '''
</body>
</html>
'''

with open(r'c:\Users\leahliu\WorkBuddy\20260525175100\schedule_extract_report.html', 'w', encoding='utf-8') as f:
    f.write(html)

print(f"Done! Total: {len(white_rows)} mainline, {len(pink_rows)} feeder, {connection_count} connections found")
print(f"Report saved to: schedule_extract_report.html")
