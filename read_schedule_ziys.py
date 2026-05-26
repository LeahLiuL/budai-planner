from openpyxl import load_workbook
from datetime import datetime, timedelta

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

pkg_col = col_map.get('PKG', 21)
pkge_col = col_map.get('PKG EB', 29)
teu_col = col_map.get('EFF. TEUS', 12)
vessel_col = col_map.get('CUL VESSELS', 3)
week_col = col_map.get('WEEK', 2)
service_col = col_map.get('SERVICES', 5)

# Get all destination port columns
port_cols = {}
for key in ['SOK', 'JED', 'MUN', 'NGB', 'KHI/AQJ (RES)', 'TAO', 'TXG', 'NAS', 'SHA', 'SHK', 'XMN']:
    if key in col_map:
        port_cols[key] = col_map[key]

remarks_col = col_map.get('REMARKS', 34)

def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    s = str(val).strip()
    for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S']:
        try: return datetime.strptime(s[:10], '%Y-%m-%d').date()
        except: pass
    return None

def is_pink(cell):
    fill = cell.fill
    try:
        rgb = str(fill.fgColor.rgb) if fill and fill.fgColor and fill.fgColor.rgb else '00000000'
    except:
        rgb = '00000000'
    pink_tags = ['DAF2D0', 'FBE2D5', 'CAEDFB']
    return any(t in rgb.upper() for t in pink_tags) or (rgb.startswith('FF') and len(rgb) == 8 and rgb[2:] != '000000')

# Find ZHI YING HE SHUN rows with ETA PKG around Jun 7
print("=== Looking for ZHI YING HE SHUN with PKG ~ Jun 7 ===\n")

target_vessel = None
target_eta_pkg = None

for row in range(header_row + 1, sheet.max_row + 1):
    vessel_cell = sheet.cell(row=row, column=vessel_col)
    vessel = str(vessel_cell.value or '').strip()
    
    if 'ZHI YING' in vessel.upper() or 'ZYHS' in vessel.upper():
        pkg_val = sheet.cell(row=row, column=pkg_col).value if pkg_col else None
        eta = parse_date(pkg_val)
        svc = sheet.cell(row=row, column=service_col).value
        teu = sheet.cell(row=row, column=teu_col).value
        pink = is_pink(vessel_cell)
        
        print(f"R{row}: {vessel} | SVC={svc} | PKG={eta} | TEU={teu} | PINK={pink}")
        
        # Target: the one with ETA PKG = 2026-06-07
        if eta and eta >= datetime(2026, 6, 1).date() and eta <= datetime(2026, 6, 15).date():
            target_vessel = vessel
            target_eta_pkg = eta
            print(f"  >>> TARGET FOUND! Row {row}, ETA PKG = {eta}")

print(f"\n=== Target: {target_vessel} @ PKG {target_eta_pkg} ===\n")

if not target_eta_pkg:
    print("ERROR: Could not find ZHI YING HE SHUN with PKG date near Jun 7")
    exit()

# Now collect ALL feeder vessels (pink) with ETD PKG after target + 1 day
cutoff = target_eta_pkg + timedelta(days=0)  # show all from this date onward
min_window = timedelta(days=1)  # minimum 1-day operation window

feeders_after = []
all_feeders = []

for row in range(header_row + 1, sheet.max_row + 1):
    vessel_cell = sheet.cell(row=row, column=vessel_col)
    vessel = str(vessel_cell.value or '').strip()
    if not vessel:
        continue
    
    pink = is_pink(vessel_cell)
    if not pink:
        continue
    
    pkg_val = sheet.cell(row=row, column=pkg_col).value if pkg_col else None
    etd = parse_date(pkg_val)
    pkge_val = sheet.cell(row=row, column=pkge_col).value if pkge_col else None
    etd_eb = parse_date(pkge_val)
    
    svc = sheet.cell(row=row, column=service_col).value
    teu = sheet.cell(row=row, column=teu_col).value
    week = sheet.cell(row=row, column=week_col).value
    remarks = sheet.cell(row=row, column=remarks_col).value if remarks_col else None
    
    # Destination ports
    dests = {}
    for pn, pc in port_cols.items():
        d = parse_date(sheet.cell(row=row, column=pc).value)
        if d:
            dests[pn] = d
    
    feeder_info = {
        'row': row,
        'vessel': vessel,
        'service': svc,
        'week': week,
        'teu': teu,
        'etd_pkg': etd,
        'etd_pkge': etd_eb,
        'dests': dests,
        'remarks': str(remarks)[:100] if remarks else ''
    }
    
    all_feeders.append(feeder_info)
    
    # Only include feeders ETD PKG on or after cutoff
    if etd and etd >= target_eta_pkg:
        wait_days = (etd - target_eta_pkg).days
        feeder_info['wait_days'] = wait_days
        feeder_info['can_connect'] = wait_days >= min_window.days
        feeders_after.append(feeder_info)

# Sort by waiting time
feeders_after.sort(key=lambda x: x.get('wait_days', 999))

print(f"Found {len(all_feeders)} total pink feeders")
print(f"Feeders with ETD PKG >= {target_eta_pkg}: {len(feeders_after)}")
print(f"Can connect (>=1day): {sum(1 for f in feeders_after if f.get('can_connect'))}")

# Build HTML
def fd(d):
    return d.strftime('%d-%b') if d else ''

html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>ZYHS 布袋船衔接计划 - PKG 7-Jun</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Segoe UI","Microsoft YaHei",sans-serif;background:#f0f2f5;padding:20px;color:#333}
h1{font-size:18px;color:#1a73e8;margin-bottom:4px}
.sub{color:#666;font-size:13px;margin-bottom:16px}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:16px;overflow:hidden}
.card-hd{background:linear-gradient(135deg,#1a73e8,#4285f4);color:#fff;padding:14px 18px;font-size:15px;font-weight:bold;display:flex;align-items:center;gap:10px}
.card-hd .badge{background:rgba(255,255,255,.25);padding:3px 12px;border-radius:15px;font-size:12px;font-weight:normal}
.info-bar{display:flex;gap:24px;padding:16px 18px;background:#f8f9fa;border-bottom:1px solid #eee;flex-wrap:wrap}
.info-item{text-align:center}
.info-num{font-size:22px;font-weight:bold;color:#1a73e8}
.info-lbl{font-size:11px;color:#888;margin-top:2px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f5f7fa;padding:10px 12px;text-align:left;font-weight:600;border-bottom:2px solid #e0e0e0;white-space:nowrap}
td{padding:9px 12px;border-bottom:1px solid #f0f0f0;white-space:nowrap}
tr:hover td{background:#fafbff}
.wait-1{color:#1565c0;font-weight:bold;font-size:14px}
.wait-2{color:#00838f;font-weight:bold}
.wait-3{color:#2e7d32;font-weight:bold}
.wait-mid{color:#ef6c00;font-weight:bold}
.wait-long{color:#c62828}
.status-ok{background:#e8f5e9;color:#2e7d32;padding:2px 8px;border-radius:4px;font-size:11px}
.status-warn{background:#fff3e0;color:#e65100;padding:2px 8px;border-radius:4px;font-size:11px}
.status-no{background:#ffebee;color:#c62828;padding:2px 8px;border-radius:4px;font-size:11px}
.teu-badge{background:#e3f2fd;color:#1565c0;padding:2px 8px;border-radius:4px;font-weight:bold;font-family:monospace}
.dest-tag{background:#f3e5f5;color:#7b1fa2;padding:2px 6px;border-radius:3px;font-size:11px;margin-left:4px}
.remark-txt{color:#999;font-style:italic;font-size:11px;max-width:250px;overflow:hidden;text-overflow:ellipsis;display:inline-block}
.timeline-dot{width:10px;height:10px;border-radius:50%;display:inline-block;margin-right:6px}
.dot-green{background:#4caf50}
.dot-yellow{background:#ffc107}
.dot-red{background:#f44336}
</style></head><body>

<div class="card">
<div class="card-hd">
  <span>🚢 ZHI YING HE SHUN — 布袋船衔接计划</span>
  <span class="badge">PKGS 到达: 07-Jun-2026</span>
</div>

<div class="info-bar">
  <div class="info-item"><div class="info-num">07-Jun</div><div class="info-lbl">ZYHS 到达 PKG</div></div>
  <div class="info-item"><div class="info-num">''' + str(len(feeders_after)) + '''</div><div class="info-lbl">可接驳船（Jun 7+）</div></div>
  <div class="info-item"><div class="info-num">''' + str(sum(1 for f in feeders_after if f.get('can_connect'))) + '''</div><div class="info-lbl">满足 ≥1天操作窗口</div></div>
  <div class="info-item"><div class="info-num">1-2天</div><div class="info-lbl">最快转接目标</div></div>
</div>

<table>
<thead><tr>
  <th>优先级</th>
  <th>接驳船 VESSEL</th>
  <th>SERVICE</th>
  <th>ETD PKG<br/>离开巴生</th>
  <th style="text-align:center">⏱️<br/>等待天数</th>
  <th style="text-align:center">状态</th>
  <th>TEU<br/>舱位</th>
  <th>目的港 ETA</th>
  <th>备注</th>
</tr></thead>
<tbody>
'''

rank = 1
for i, f in enumerate(feeders_after):
    w = f.get('wait_days', '?')
    can = f.get('can_connect', False)
    
    if isinstance(w, int):
        if w <= 2:
            wc = 'wait-1'; dot = 'dot-green'
        elif w <= 4:
            wc = 'wait-2'; dot = 'dot-yellow'
        elif w <= 7:
            wc = 'wait-mid'; dot = 'dot-yellow'
        else:
            wc = 'wait-long'; dot = 'dot-red'
        
        status = '<span class="status-ok">✅ 可接</span>' if can else f'<span class="status-no">❌ 不足1天</span>'
        wait_str = f'<td class="{wc}" style="text-align:center">{w} 天</td>'
    else:
        wc = ''; dot = 'dot-red'
        status = '<span class="status-warn">⚠️ 待确认</span>'
        wait_str = f'<td class="wait-long" style="text-align:center">?</td>'
    
    # Dest ports as tags
    dest_str = ''
    for pn, pd in sorted(f['dests'].items()):
        short_name = pn.split('(')[0].strip().replace('/AQJ (RES)', '')
        dest_str += f'<span class="dest-tag">{short_name} {pd.strftime("%d/%m")}</span>'
    
    teu_str = str(f['teu']) if f['teu'] else 'TBC'
    
    rank_color = "#1a73e8" if rank <= 3 else "#999"
    html += '<tr>'
    html += f'<td style="text-align:center;font-weight:bold;color:{rank_color}">{rank}</td>'
    html += f'<td><strong>{f["vessel"]}</strong></td>'
    html += f'<td>{f["service"] or ""}</td>'
    html += f'<td style="font-family:monospace">{fd(f["etd_pkg"])}</td>'
    html += wait_str
    html += f'<td style="text-align:center">{status}</td>'
    html += f'<td style="text-align:center"><span class="teu-badge">{teu_str}</span></td>'
    html += f'<td>{dest_str}</td>'
    html += f'<td><span class="remark-txt">{f["remarks"]}</span></td>'
    html += '</tr>\n'
    rank += 1

html += '''</tbody></table>

<div style="padding:16px 18px;background:#fffbf0;border-top:2px solid #ffd54f;font-size:13px;">
<strong>💡 排船建议：</strong>
'''
# Generate recommendation
best_1_2d = [f for f in feeders_after if isinstance(f.get('wait_days'), int) and 1 <= f['wait_days'] <= 2]
best_3_5d = [f for f in feeders_after if isinstance(f.get('wait_days'), int) and 3 <= f['wait_days'] <= 5]

if best_1_2d:
    html += f"<br/><strong>🟢 最优快转（1-2天等待）：</strong> 共 {len(best_1_2d)} 条接驳船可选"
    for b in best_1_2d[:3]:
        html += f"<br/>&nbsp;&nbsp;• <strong>{b['vessel']}</strong> — ETD PKG {fd(b['etd_pkg'])}，等 {b['wait_days']}天，舱位 {b['teu'] or 'TBC'}TEU"

if best_3_5d:
    html += f"<br/><br/><strong>🟡 正常转接（3-5天等待）：</strong> 共 {len(best_3_5d)} 条备选"
    for b in best_3_5d[:3]:
        html += f"<br/>&nbsp;&nbsp;• <strong>{b['vessel']}</strong> — ETD PKG {fd(b['etd_pkg'])}，等 {b['wait_days']}天，舱位 {b['teu'] or 'TBC'}TEU"

long_wait = [f for f in feeders_after if isinstance(f.get('wait_days'), int) and f['wait_days'] > 5]
if long_wait:
    html += f"<br/><br/><strong>⚪ 长期备选（>5天）：</strong> {len(long_wait)} 条（仅作应急储备）"

html += '</div></div></body></html>'

outpath = r'c:\Users\leahliu\WorkBuddy\20260525175100\feeder_plan_ZYHS.html'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\nReport saved to: {outpath}")
