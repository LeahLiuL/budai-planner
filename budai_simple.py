from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=7, column=c).value
    if h: col_map[h.strip().upper()] = c

pkg_c = col_map.get('PKG', 21)
teu_c = col_map.get('EFF. TEUS', 12)
ves_c = col_map.get('CUL VESSELS', 3)
svc_c = col_map.get('SERVICES', 5)

def to_dt(val):
    """Convert cell value to datetime, return raw str if fail"""
    if val is None: return None, ''
    if isinstance(val, datetime): return val, val.strftime('%m/%d')
    s = str(val).strip()
    for f in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(s[:19] if len(s)>=19 else s[:10], f)
            return dt, dt.strftime('%m/%d')
        except: pass
    return None, s

def get_rgb(cell):
    try:
        r = str(cell.fill.fgColor.rgb).upper()
        return r if len(r)==8 else '00000000'
    except: return '00000000'

def color_type(rgb):
    if rgb == '00000000': return 'W'
    if 'DAF2D0' in rgb: return 'P'
    if 'FBE2D5' in rgb or 'BF' in rgb[2:] or 'CB' in rgb[2:]: return 'F'
    return '?'

# Collect rows 51-79
rows = []
for row in range(51, 80):
    v = str(sheet.cell(row, ves_c).value or '').strip()
    if not v: continue
    
    cell = sheet.cell(row, ves_c)
    rgb = get_rgb(cell); ct = color_type(rgb)
    
    pkg_raw = sheet.cell(row, pkg_c).value
    pkg_dt, pkg_str = to_dt(pkg_raw)
    teu_val = sheet.cell(row, teu_c).value
    svc = str(sheet.cell(row, svc_c).value or '')
    
    teu_num = int(teu_val) if str(teu_val).isdigit() else (teu_val or 0)
    
    rows.append({
        'row': row, 'name': v, 'type': ct,
        'pkg_dt': pkg_dt, 'pkg': pkg_str,
        'teu': teu_num, 'svc': svc
    })

white = [r for r in rows if r['type'] == 'W']
feeders = [r for r in rows if r['type'] == 'F']
pink = [r for r in rows if r['type'] == 'P']

# Match white -> feeder
assignments = []  # list of dicts
for w in white:
    we = w['pkg_dt']
    if not we: continue
    cargo = w['teu']
    remain = cargo
    
    opts = []
    for fi, f in enumerate(feeders):
        fe = f['pkg_dt']
        if not fe: continue
        days = (fe - we).days
        if days >= 1: opts.append((fi, days))
    opts.sort(key=lambda x: x[1])
    
    for fi, days in opts:
        if remain <= 0: break
        f = feeders[fi]
        cap = f['teu']
        used = sum(a['ateu'] for a in assignments if a['feeder_row'] == f['row'])
        avail = cap - used
        if avail <= 0: continue
        
        take = min(remain, avail)
        assignments.append({
            'w': w, 'f': f,
            'days': days, 'ateu': take,
            'feeder_row': f['row'], 'w_row': w['row']
        })
        remain -= take

# Calc unassigned per white ship
unas = {}
for w in white:
    done = sum(a['ateu'] for a in assignments if a['w_row'] == w['row'])
    unas[w['row']] = max(0, w['teu'] - done)

total_cargo = sum(w['teu'] for w in white)
total_done = sum(a['ateu'] for a in assignments)
total_unas = total_cargo - total_done

# Build HTML
L = []; A = L.append

A('<!DOCTYPE html><html><head><meta charset="UTF-8"><title>布袋船计划</title>')
A('''<style>
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:#f5f6fa;padding:24px;max-width:1100px;margin:auto;color:#222}
h1{font-size:20px;margin-bottom:2px}.sub{color:#888;font-size:13px;margin-bottom:16px}
.card{background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:14px;overflow:hidden}
.hd{padding:10px 14px;font-weight:bold;font-size:15px;border-bottom:1px solid #eee}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#fafafa;padding:7px 10px;text-align:left;font-weight:600;border-bottom:2px solid #ddd}
td{padding:6px 10px;border-bottom:1px solid #f2f2f2}tr:hover td{background:#f9faff}
.ok{color:#2e7d32;font-weight:600}.warn{color:#e65100;font-weight:600}.bad{color:#c62828;font-weight:600}
.unas{background:#ffebee;color:#c62828;padding:2px 8px;border-radius:4px;font-weight:bold}
.tip{background:#e8f5e9;padding:10px 14px;font-size:13px;border-radius:6px;margin-top:12px;line-height:1.6}
.warn-box{background:#fff3e0;padding:10px 14px;border-radius:6px;font-size:13px;line-height:1.6}
.stats{display:flex;gap:16px;padding:14px 18px}.stat{text-align:center}
.stat .n{font-size:22px;font-weight:bold;color:#1565c0}.stat .l{font-size:11px;color:#666}
</style></head><body>''')

A('<h1>🚢 布袋船排计划</h1>')
A('<div class="sub">R51~R79 | ZHI YING HE SHUN 06/07 之后 | 2026-05-25</div>')

# Stats bar
A(f'<div class="card"><div class="stats">')
for label, val, clr in [('干线船',len(white),'1565c0'),('布袋船',len(feeders),'e65100'),
                         ('已安排TEU',total_done,'2e7d32'),('未安排TEU',total_unas,'c62828')]:
    A(f'<div class="stat"><div class="n" style="color:{clr}">{val}</div><div class="l">{label}</div></div>')
A('</div></div>')

# Plan table
A('<div class="card"><div class="hd">📋 衔接计划（按到港时间排序）</div>')
A('<table><thead><tr><th>#</th><th>干线船</th><th>到巴生</th><th>货量</th><th>→ 布袋船接</th><th>离巴生</th><th>等天</th><th>安排TEU</th><th>状态</th></tr></thead><tbody>')

plan = sorted(assignments, key=lambda x: x['w']['pkg_dt'] or datetime.max)

for i, a in enumerate(plan, 1):
    w=a['w']; f=a['f']; d=a['days']; t=a['ateu']
    cl = 'ok' if d<=5 else ('warn' if d<=10 else 'bad')
    st = '✅快转' if d<=5 else ('⚠️正常' if d<=10 else '🔴较长')
    A(f"<tr><td>{i}</td><td><b>{w['name']}</b></td><td>{w['pkg']}</td><td>{w['teu']}</td>")
    A(f"<td>{f['name']}</td><td>{f['pkg']}</td><td class='{cl}'>{d}</td><td>{t}</td><td class='{cl}'>{st}</td></tr>")

# Unassigned rows
shown = set()
for a in assignments:
    wr = a['w_row']
    if wr in shown: continue
    u = unas.get(wr, 0)
    if u > 0:
        shown.add(wr)
        w = a['w']
        A(f"<tr style='background:#fff8e1'><td></td><td><b>{w['name']}</b></td><td>{w['pkg']}</td><td>{w['teu']}</td>")
        A("<td colspan='4' style='color:#999'>— 无可用布袋船衔接 —</td>")
        A(f"<td><span class='unas'>{u} TEU 未安排</span></td></tr>")

A('</tbody></table></div>')

# Feeder utilization
f_used = {f['row']:0 for f in feeders}
for a in assignments: f_used[a['feeder_row']] += a['ateu']

A("<div class='card'><div class='hd'>📦 布袋船 BSA 利用率</div>")
A('<table><thead><tr><th>布袋船</th><th>离巴生</th><th>容量(TEU)</th><th>已用</th><th>剩余</th><th>利用率</th></tr></thead><tbody>')
for f in feeders:
    u = f_used[f['row']]; cap = f['teu']; left = cap-u; pct = round(u/cap*100) if cap else 0
    bc = '#4caf50' if pct>=80 else ('#ff9800' if pct>=30 else '#ccc')
    bar = f"<div style='background:#eee;height:8px;border-radius:4px;width:120px'><div style='background:{bc};height:100%;width:{max(pct,3)}%;border-radius:4px'></div></div>"
    A(f"<tr><td><b>{f['name']}</b></td><td>{f['pkg']}</td><td>{cap}</td><td><b>{u}</b></td><td>{left}</td><td>{pct}% {bar}</td></tr>")
A('</tbody></table></div>')

# Recommendations
if total_unas > 0:
    A("<div class='card'><div class='hd'>💡 加船建议</div>")
    A(f"<div class='warn-box'><b>⚠️ 共 {total_unas} TEU 货物未安排衔接，建议：</b><br><br>")
    
    unas_list = [(w, unas[w['row']]) for w in white if unas.get(w['row'],0)>0]
    unas_list.sort(key=lambda x: x[0]['pkg_dt'] or datetime.max)
    
    for idx, (w, need) in enumerate(unas_list, 1):
        suggest_teus = min(need, 1700)
        A(f"{idx}. <b>{w['pkg']} 前后</b> — {w['name']} 有 <b>{need} TEU</b> 未接<br>")
        A(f"   → 建议在 {w['pkg']} 后加一条 ≥{suggest_teus}TEU 布袋船<br><br>")
    
    A("</div></div>")

# Legend
A("<div class='card tip'>")
A("📌 <b>说明：</b> 白底=国内→巴生 | 橙色=巴生→红海布袋船 | 操作窗口≥1天 | 绿色≤5天为快转")
A("</div>")

A("</body></html>")

out = r'c:\Users\leahliu\WorkBuddy\20260525175100\budai_plan.html'
with open(out, 'w', encoding='utf-8') as f:
    f.write('\n'.join(L))

print(f"Cargo={total_cargo}, Assigned={total_done}, Unassigned={total_unas}, Assigns={len(assignments)}")
print("Output:", out)
