from openpyxl import load_workbook
from datetime import datetime, timedelta

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h: col_map[h.strip().upper()] = c

pkg_c   = col_map.get('PKG',21)
pkge_c  = col_map.get('PKG EB',29)
teu_c   = col_map.get('EFF. TEUS',12)
vol_wb_c= col_map.get('DISCHARGE VOL',30)
vessel_c= col_map.get('CUL VESSELS',3)
week_c  = col_map.get('WEEK',2)
svc_c   = col_map.get('SERVICES',5)
voyage_c= col_map.get('VOYAGE .NO',4)
rmk_c   = col_map.get('REMARKS',34)

dest_cols={}
for p in ['SOK','JED','MUN','NGB']:
    if p in col_map: dest_cols[p]=col_map[p]

def fmt(d):
    if d is None:return ''
    if isinstance(d,datetime):return d.strftime('%Y-%m-%d')
    s=str(d)[:10];return s if s else ''
def get_rgb(cell):
    try:
        r=str(cell.fill.fgColor.rgb).upper()
        return r if len(r)==8 else '00000000'
    except:return '00000000'

# Color classification (confirmed by user)
def classify(rgb):
    if rgb=='00000000':return ('WHITE','#e8eaed','白底·国内→巴生')
    if 'DAF2D0' in rgb: return ('PINK_DIRECT','#fce4ec','粉底·国内→红海直达')
    if any(x in rgb for x in ['FBE2D5']): return ('FEEDER_ORANGE','#fff3e0','橙底·巴生→红海布袋')
    # catch-all for other non-white colors
    if rgb.startswith('FF') and rgb[2:]!='000000': return ('FEEDER_OTHER','#fff8e1','其他色·接驳船')
    return ('OTHER','#eee','未知')

def parse_dt(s):
    if not s:return None
    try:return datetime.strptime(s,'%Y-%m-%d')
    except:return None

# === Collect R51-R79 ===
all_rows=[]
for row in range(51,80):
    vc=sheet.cell(row=row,column=vessel_c); v=str(vc.value or '').strip()
    if not v:continue
    rgb=get_rgb(vc); ctype,css,lab=classify(rgb)
    pkg=sheet.cell(row=row,column=pkg_c).value; pkge=sheet.cell(row=row,column=pkge_c).value
    teu=sheet.cell(row=row,column=teu_c).value
    vol_wb=sheet.cell(row=row,column=vol_wb_c).value
    wk=sheet.cell(row=row,column=week_c).value; svc=sheet.cell(row=row,column=svc_c).value
    voy=sheet.cell(row=row,column=voyage_c).value; rmk=sheet.cell(row=row,column=rmk_c).value
    wk=sheet.cell(row=row,column=week_c).value; svc=sheet.cell(row=row,column=svc_c).value
    voy=sheet.cell(row=row,column=voyage_c).value; rmk=sheet.cell(row=row,column=rmk_c).value
    dst={}; dest_strs=[]
    for pn,pc in dest_cols.items():
        val=sheet.cell(row=row,column=pc).value
        if val:dst[pn]=fmt(val);dest_strs.append(pn+':'+fmt(val))
    
    all_rows.append({
        'row':row,'vessel':v,'week':wk,'svc':str(svc)if svc else'','voy':str(voy)if voy else'',
        'pkg':fmt(pkg),'pkge':fmt(pkge),'teu':teu,
        'rgb':rgb,'ctype':ctype,'label':lab,'color':css,
        'rmk':(str(rmk)[:80]if rmk else''),'dst':dst,'dest_str':', '.join(dest_strs),
        'vol_wb':vol_wb
    })

white=[r for r in all_rows if r['ctype']=='WHITE']
feeder=[r for r in all_rows if r['ctype'].startswith('FEEDER')]
pink_direct=[r for r in all_rows if r['ctype']=='PINK_DIRECT']
other=[r for r in all_rows if r['ctype'] not in('WHITE','FEEDER_ORANGE','FEEDER_OTHER','PINK_DIRECT')]

print(f"Rows: {len(all_rows)} | White={len(white)} | Feeder={len(feeder)} | PinkDirect={len(pink_direct)} | Other={len(other)}")

# === Determine cargo volume per white ship ===
MIN_WAIT_DAYS = 1

for w in white:
    teu_val=w['teu']
    vol_val=w['vol_wb']
    # Use DISCHARGE VOL (WB PKG CARGO VOLUME) first, fallback to TEU
    if vol_val is not None:
        try: w['cargo_teu']=int(float(vol_val))
        except: w['cargo_teu']=int(teu_val) if str(teu_val).isdigit() else 1100
    elif teu_val is not None:
        try: w['cargo_teu']=int(teu_val)
        except: w['cargo_teu']=1100
    else:
        w['cargo_teu']=1100
    w['assigned']=0
    w['unassigned']=w['cargo_teu']

# Set BSA capacity for feeders
for f in feeder:
    teu_val=f['teu']
    if teu_val is not None:
        try: f['bsa_cap']=int(teu_val)
        except: f['bsa_cap']=1000
    else:
        f['bsa_cap']=1000
    f['bsa_used']=0

# === Build all valid connections ===
connections=[]
for mw in white:
    eta=parse_dt(mw['pkg'])
    if not eta: continue
    matches=[]
    for mp in feeder:
        etd=parse_dt(mp['pkg'])
        if etd and (etd-eta).days>=MIN_WAIT_DAYS:
            delta=(etd-eta).days
            cls='good' if delta<=3 else ('warn' if delta<=5 else 'bad')
            matches.append({'f':mp,'delta':delta,'cls':cls})
    matches.sort(key=lambda x:x['delta'])
    connections.append({'main':mw,'matches':matches})
connections.sort(key=lambda c: parse_dt(c['main']['pkg'])or datetime.max)

# === Greedy assignment: maximize BSA fill + minimize wait ===
final_assignments=[]

for conn in connections:
    mw=conn['main']; remaining=mw['cargo_teu']-mw['assigned']
    if remaining<=0: continue
    eta=parse_dt(mw['pkg'])

    for m in conn['matches']:
        if remaining<=0: break
        mf=m['f']; cap=mf['bsa_cap']
        avail=cap-mf['bsa_used']
        if avail<=0: continue
        assign=min(remaining,avail)
        
        mf['bsa_used']+=assign
        mw['assigned']+=assign
        mw['unassigned']-=assign
        
        final_assignments.append({
            'mainline':mw,'feeder':mf,
            'wait_days':m['delta'],'status_cls':m['cls'],
            'assigned_teu':assign,'bsa_cap':cap,
            'bsa_fill_pct':mf['bsa_used']/cap*100
        })

# === Generate recommendations for unassigned cargo ===
recommendations=[]
for w in white:
    if w['unassigned']>0:
        eta=parse_dt(w['pkg'])
        rec_date=(eta+timedelta(days=3)).strftime('%Y-%m-%d') if eta else 'TBC'
        recommendations.append({
            'mainline':w,'unassigned_teus':w['unassigned'],
            'rec_etd_pkg':rec_date,'rec_bsa_needed':w['unassigned'],
            'reason':f"到PKG后{rec_date}左右需额外调布袋船"
        })

# === Build HTML with string list (no nested f-string issues) ===
L=[];A=L.append
def esc(s):return str(s).replace('<','&lt;').replace('>','&gt;')

A('''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>布袋船排计划 — R51~R79</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Segoe UI","Microsoft YaHei",sans-serif;background:#f0f2f5;padding:20px;color:#222}
h1{font-size:22px;margin-bottom:4px;color:#1565c0}
.sub{color:#666;font-size:13px;margin-bottom:18px}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 12px rgba(0,0,0,.08);margin-bottom:18px;overflow:hidden}
.hd{padding:14px 18px;font-weight:700;font-size:15px;display:flex;align-items:center;gap:10px;border-bottom:1 solid #e8e8e8}
.hd .badge-pill{display:inline-block;padding:2px 12px;border-radius:14px;font-size:12px;font-weight:400}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#f7f8fa;padding:8px 10px;text-align:left;font-weight:600;border-bottom:2px solid #dde1e6}
td{padding:6px 10px;border-bottom:1 solid #f0f0f0}
tr:hover td{background:#f5f8ff}
.tag{display:inline-block;padding:2px 9px;border-radius:10px;font-size:11px;font-weight:500}
.tg-w{background:#e8eaed;color:#333}tg-p{background:#fce4ec;color:#a61b4d}tg-f{background:#fff3e0;color:#c45600}tg-o{background:#fff8e1;color:#7a6000}tg-x{background:#eee;color:#666}
.date-f{font-family:"Courier New",monospace;font-size:11.5px}
.teu-n{font-weight:700;text-align:right}.good{color:#1b5e20;font-weight:700}.warn{color:#bf360c;font-weight:700}.bad{color:#b71c1c;font-weight:700}
.rmks{color:#777;font-style:italic;font-size:11px;max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:normal}
.mc{border:1 solid #e0e0e0;border-radius:8px;margin-bottom:14px;overflow:hidden}
.mh{background:#e3f2fd;padding:10px 14px;display:flex;align-items:center;gap:12px;font-weight:700;font-size:14px}
.mb{padding:10px 14px}
.mr{display:flex;align-items:center;padding:5px 0;border-bottom:1 dotted #ececec;font-size:12px}
.mr:last-child{border:none}
.summ{display:flex;gap:28px;padding:16px 20px;flex-wrap:wrap}
.si .num{font-size:30px;font-weight:800;color:#1565c0}
.si .lab{font-size:12px;color:#666;margin-top:2px}
.lb{padding:10px 18px;display:flex;gap:22px;font-size:12px;border-top:1 solid #eee;background:#fafafa;flex-wrap:wrap}
.cb{width:20px;height:20px;border-radius:4px;border:1 solid #ccc;display:inline-block;vertical-align:middle}
/* Unassigned highlight */
.unassigned-row{background:#fffde7!important}
.unassigned-row:hover td{background:#fff9c4!important}
.warn-box{border:2 solid #ffc107;border-radius:8px;background:#fff8e1;padding:14px 18px;margin-bottom:14px}
.warn-title{font-weight:700;font-size:14px;color:#e65100;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.rec-box{border:2 solid #4caf50;border-radius:8px;background:#f1f8e9;padding:14px 18px;margin-bottom:14px}
.rec-title{font-weight:700;font-size:14px;color:#2e7d32;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.rec-item{padding:6px 0;border-bottom:1 dashed #c5e1a5;font-size:13px}
.rec-item:last-child{border:none}
.progress{height:18px;background:#eee;border-radius:9px;overflow:hidden;position:relative}
.progress-bar{height:100%;border-radius:9px;transition:width .3s}
.progress-text{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);font-size:10px;font-weight:700;z-index:1}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;padding:16px 20px}
.kpi-card{background:#f8f9ff;border:1 solid #e0e8ff;border-radius:8px;padding:12px;text-align:center}
.kpi-val{font-size:26px;font-weight:800;color:#1565c0}
.kpi-label{font-size:11px;color:#555;margin-top:2px}
.status-dot{width:8px;height:8px;border-radius:50%;display:inline-block}
.dot-green{background:#4caf50}.dot-red{background:#f44336}.dot-yellow{background:#ff9800}
</style></head><body>''')

# ===== HEADER =====
A('<h1>🚢 布袋船（Budai）衔接计划</h1>')
A('<div class="sub">REX & FEEDER SCHEDULE · Sheet: <b>25_May</b> · 范围: R51 ~ R79 · 生成时间: '+datetime.now().strftime('%Y-%m-%d %H:%M')+'</div>')

# ===== KPI SUMMARY =====
total_cargo=sum(w['cargo_teu']for w in white)
total_assigned=sum(w['assigned']for w in white)
total_unassigned=sum(w['unassigned']for w in white)
total_bsa=sum(f['bsa_cap']for f in feeder)
total_bsa_used=sum(f['bsa_used']for f in feeder)

avg_wait=0
wt_sum=0
for fa in final_assignments:
    avg_wait+=fa['wait_days']*fa['assigned_teu']
    wt_sum+=fa['assigned_teu']
avg_wait=round(avg_wait/wt_sum,1)if wt_sum else 0

A('<div class="card">')
A('<div class="hd">📊 计划概览 KPI</div>')
A('<div class="kpi-grid">')
A(f"<div class='kpi-card'><div class='kpi-val'>{len(white)}</div><div class='kpi-label'>干线船（国内→巴生）</div></div>")
A(f"<div class='kpi-card'><div class='kpi-val'>{len(feeder)}</div><div class='kpi-label'>布袋船（巴生→红海）</div></div>")
A(f"<div class='kpi-card'><div class='kpi-val'>{total_cargo:,}</div><div class='kpi-label'>总到巴生货量 (TEU)</div></div>")
pct_assigned=round(total_assigned/total_cargo*100)if total_cargo else 0
A(f"<div class='kpi-card'><div class='kpi-val'>{total_assigned:,}</div><div class='kpi-label'>已安排 ({pct_assigned}%)</div></div>")
A(f"<div class='kpi-card' style='border-color:#ffebee;background:#fffde7'><div class='kpi-val' style='color:#c62828'>{total_unassigned:,}</div><div class='kpi-label'>⚠️ 未安排</div></div>")
pct_bsa=round(total_bsa_used/total_bsa*100)if total_bsa else 0
A(f"<div class='kpi-card'><div class='kpi-val'>{pct_bsa}%</div><div class='kpi-label'>BSA 舱位利用率</div></div>")
A(f"<div class='kpi-card'><div class='kpi-val'>{avg_wait}</div><div class='kpi-label'>加权平均等待(天)</div></div>")
A('</div></div>')

# Legend
A('<div class="card"><div class="lb">')
A("<span><span class='cb' style='background:#fff'></span> 白底 = 干线船 国内→巴生（卸货）</span>")
A("<span><span class='cb' style='background:#DAF2D0'></span> 粉底 = 直航 国内→红海（不经过巴生）</span>")
A("<span><span class='cb' style='background:#FBE2D5'></span> 橙底 = 布袋船 巴生→红海（装货）</span>")
A("<span style='margin-left:auto'><span class='status-dot dot-green'></span> ≤3天快转 &nbsp; <span class='status-dot dot-yellow'></span> 4-5天 &nbsp; <span class='status-dot dot-red'></span> >5天较慢</span>")
A('</div></div>')

# ===== SECTION 1: Full schedule table =====
A("<div class='card'><div class='hd'>📋 全部船期表 R51~R79（含颜色分类验证）</div>")
A('<div style="overflow-x:auto"><table><thead><tr>')
cols=['R#','RGB','分类','VESSEL','VOYAGE','SVC','TEU','CARGO(TEU)','ETA_PKG','PKGE','SOK','JED','MUN','NGB','状态','REMARKS']
for c in cols:A(f'<th>{c}</th>')
A('</tr></thead><tbody>')

for r in all_rows:
    bg={'WHITE':'w','PINK_DIRECT':'p','FEEDER_ORANGE':'f','FEEDER_OTHER':'o','OTHER':'x'}.get(r['ctype'],'x')
    status=''
    if r['ctype']=='WHITE':
        if r.get('unassigned',0)>0:
            status='<span style="color:#c62828;font-weight:700">⚠ 未排完 '+str(r.get('unassigned',0))+'TEU</span>'
        elif r.get('assigned',0)>0:
            status='<span style="color:#2e7d32;font-weight:700">✅ 已安排 '+str(r.get('assigned',0))+'TEU</span>'
        row_class=' unassigned-row' if r.get('unassigned',0)>0 else''
    else:
        fill_pct=r.get('bsa_used',0)/r.get('bsa_cap',1)*100 if r.get('bsa_cap',0)>0 else 0
        if r['ctype'].startswith('FEEDER'):
            status=f'{r.get("bsa_used",0)}/{r.get("bsa_cap","?")} TEU ({fill_pct:.0f}%)'
        row_class=''
    
    A(f"<tr{row_class}><td>{r['row']}</td><td style='font-size:10px;font-family:monospace'>{r['rgb']}</td>")
    A(f"<td><span class='tag tg-{bg}'>{r['label']}</span></td>")
    A(f"<td><strong>{esc(r['vessel'])}</strong></td><td>{esc(r['voy'])}</td><td>{esc(r['svc'])}</td><td class='teu-n'>{r['teu']}</td>")
    if r['ctype']=='WHITE':
        A(f"<td class='teu-n'><b>{r.get('cargo_teu','-')}</b></td>")
    else:
        A(f"<td class='teu-n'>-</td>")
    A(f"<td class='date-f'>{r['pkg']}</td><td class='date-f'>{r['pkge']}</td>")
    A(f"<td class='date-f'>{r['dst'].get('SOK','')}</td><td class='date-f'>{r['dst'].get('JED','')}</td>")
    A(f"<td class='date-f'>{r['dst'].get('MUN','')}</td><td class='date-f'>{r['dst'].get('NGB','')}</td>")
    A(f"<td>{status}</td><td class='rmks'>{esc(r['rmk'])}</td></tr>")

A('</tbody></table></div></div>')

# ===== SECTION 2: Connection plan per mainline ship =====
A("<div class='card'><div class='hd'>🔗 衔接计划详情（按 ETA PKG 时间排序）</div>")
A('<div style="padding:14px 18px"><p style="font-size:13px;color:#555">每条白底干线船到达巴生后的布袋船分配方案。<b style="color:#c62828">黄色高亮行=有未安排货量</b></p>')

for ci,c in enumerate(connections):
    m=c['main'];ms=c['matches']
    
    # Status icon
    if m['unassigned']>0:icon_status='⚠️';st_color='#c62828';st_text=f'未安排 {m["unassigned"]}TEU'
    elif m['assigned']>0:icon_status='✅';st_color='#2e7d32';st_text=f'已安排 {m["assigned"]}TEU / 共 {m["cargo_teu"]}TEU'
    else:icon_status='❓';st_color='#666';st_text='无有效衔接'
    
    A(f"<div class='mc'>")
    A(f"<div class='mh'>")
    mcolor = m['color']
    mvessel = esc(m['vessel'])
    A(f"<span style='background:{mcolor};padding:4px 10px;border-radius:4px'>{mvessel}</span>")
    A(f"<small>ETA PKG: <code>{m['pkg']}</code></small>")
    A(f"<small>| {esc(m['svc'])} | 总货 {m['cargo_teu']}TEU</small>")
    A(f"<span style='margin-left:auto;color:{st_color};font-weight:600'>{icon_status} {st_text}</span>")
    A("</div>")  # mh
    
    if ms:
        A("<div class='mb'>")
        shown=0
        for mm in ms[:15]:
            ff=mm['f'];dd=ff['dst']
            parts=[]
            for p in ['SOK','JED','MUN','NGB']:
                if dd.get(p):parts.append(p+':'+dd[p])
            ds=', '.join(parts)if parts else '-'
            
            # Check how much assigned from this mainline to this feeder
            assign_amt=0
            for fa in final_assignments:
                if fa['mainline']==m and fa['feeder']==ff:
                    assign_amt=fa['assigned_teu'];break
            
            bar_color='#4caf50'if mm['cls']=='good'else('#fb8c00'if mm['cls']=='warn'else('#ef5350'))
            
            A("<div class='mr'>")
            A(f"<span style='width:190px'><b>{esc(ff['vessel'])}</b><small>({esc(ff['svc'])})</small></span>")
            A(f"<span class='date-f' style='width:90px'>ETD {ff['pkg']}</span>")
            A(f"<span class='{mm['cls']}' style='width:58px;text-align:center'>等{mm['delta']}天</span>")
            A(f"<span style='width:90px;text-align:right;font-weight:700;color:{bar_color}'>{assign_amt} TEU</span>"if assign_amt else "<span>-</span>")
            A(f"<span style='width:70px;text-align:right'>/{ff['bsa_cap']} TEU</span>")
            # Progress bar
            pct=ff['bsa_used']/ff['bsa_cap']*100 if ff['bsa_cap']>0 else 0
            bar_w=min(pct,100)
            A(f"<div class='progress'style='width:120px'><div class='progress-bar'style='width:{bar_w}%;background:{bar_color}'></div><span class='progress-text'>{ff['bsa_used']}/{ff['bsa_cap']}({pct:.0f}%)</span></div>")
            A(f"<span style='color:#888;margin-left:8px'>→ {ds}</span>")
            A("</div>")  # mr
            shown+=1
        if len(ms)>shown:
            A(f"<div style='font-size:11px;color:#999;padding:3px 0'>... 还有 {len(ms)-shown} 条可衔接（等待更长）</div>")
        A("</div>")  # mb
    else:
        A("<div class='mb'><p style='color:#c62828'>❌ 该船到达巴生后无可用的布袋船（无满足≥1天窗口的橙底船）</p></div>")
    A("</div>\n")  # mc

A('</div></div>')  # card

# ===== SECTION 3: Global sorted table =====
A("<div class='card'><div class='hd'>📌 衔接总表（按干线船 ETA PKG 排序）</div>")
A('<div style="overflow-x:auto"><table><thead>')
A('<tr><th>#</th><th>干线船(白)</th><th>ETA PKG</th><th>布袋船(橙)</th><th>ETD PKG</th><th>等天</th><th>安排TEU</th><th>状态</th><th>BSA容量</th><th>填充率</th><th>SVC/目的港</th></tr>')
A('</thead><tbody>')

sorted_fa=sorted(final_assignments,key=lambda x:parse_dt(x['mainline']['pkg'])or datetime.max)

for i,fa in enumerate(sorted_fa,1):
    m=fa['mainline'];f=fa['feeder']
    dd=f['dst']
    parts=[];ds='-'
    for p in ['SOK','JED','MUN','NGB']:
        if dd.get(p):parts.append(p+':'+dd[p])
    ds=', '.join(parts)if parts else '-'
    
    wcls=fa['status_cls']
    status='✅ 快转'if wcls=='good'else('⚠️ 正常'if wcls=='warn'else('🔴 较慢'))
    fill_pct=fa['bsa_fill_pct']
    
    A(f"<tr><td>{i}</td><td><b>{esc(m['vessel'])}</b></td><td class='date-f'>{m['pkg']}</td>")
    A(f"<td><b>{esc(f['vessel'])}</b></td><td class='date-f'>{f['pkg']}</td>")
    A(f"<td class='{wcls}'>{fa['wait_days']}</td><td class='teu-n'><b>{fa['assigned_teu']}</b></td>")
    A(f"<td>{status}</td><td class='teu-n'>{fa['bsa_cap']}</td><td>{fill_pct:.0f}%</td>")
    A(f"<td>{esc(f['svc'])} → {ds}</td></tr>")

A('</tbody></table></div></div>')

# ===== SECTION 4: UNASSIGNED WARNING =====
if total_unassigned > 0:
    A('<div class="warn-box">')
    A('<div class="warn-title">⚠️ 未安排货量明细 — 需要额外布袋船或调整计划</div>')
    A('<table><thead><tr><th>#</th><th>干线船</th><th>ETA PKG</th><th>总货量</th><th>已安排</th><th style="color:#c62828;font-weight:700">未安排</th><th>未安排率</th></tr></thead><tbody>')
    
    unassigned_ships=[w for w in white if w['unassigned']>0]
    unassigned_ships.sort(key=lambda w:parse_dt(w['pkg'])or datetime.max)
    
    for ui,w in enumerate(unassigned_ships,1):
        urate=round(w['unassigned']/w['cargo_teu']*100)
        A(f"<tr class='unassigned-row'><td>{ui}</td><td><b>{esc(w['vessel'])}</b></td><td class='date-f'>{w['pkg']}</td>")
        A(f"<td class='teu-n'>{w['cargo_teu']}</td><td class='teu-n'>{w['assigned']}</td><td class='teu-n' style='color:#c62828;font-weight:800'>{w['unassigned']}</td>")
        A(f"<td>{urate}%</td></tr>")
    
    A(f"</tbody></table><div style='margin-top:10px;font-size:13px'><b>合计未安排: {total_unassigned:,} TEU ({round(total_unassigned/total_cargo*100)}%)</b></div>")
    A('</div>')  # warn-box

# ===== SECTION 5: RECOMMENDATIONS =====
if recommendations:
    A('<div class="rec-box">')
    A('<div class="rec-title">💡 加船建议 — 在以下时间点增加布袋船可最大化覆盖未安排货物</div>')
    
    # Group recommendations by date window
    rec_groups={}
    for rec in recommendations:
        key=rec['rec_etd_pkg']
        if key not in rec_groups:rec_groups[key]=[]
        rec_groups[key].append(rec)
    
    ri=0
    for rd in sorted(rec_groups.keys()):
        grp=rec_groups[rd];ri+=1
        total_need=sum(r['rec_bsa_needed']for r in grp)
        ships=', '.join([r['mainline']['vessel']for r in grp])
        
        A(f"<div class='rec-item'>")
        A(f"<b>建议 #{ri}</b>: 在 <code>{rd}</code> 左右 ETD PKG 安排一条 <b>≥{total_need} TEU</b> 的布袋船")
        A(f"<br><small style='color:#555'>覆盖船只: {ships} | 可解决 {total_need} TEU 未安排货物</small>")
        A(f"</div>")
    
    A('</div>')  # rec-box

# ===== SECTION 6: Feeder utilization =====
A("<div class='card'><div class='hd'>📦 各布袋船 BSA 利用情况</div>")
A('<div style="overflow-x:auto"><table><thead>')
A('<tr><th>#</th><th>布袋船(橙)</th><th>SVC</th><th>ETD PKG</th><th>BSA容量</th><th>已用</th><th>剩余</th><th>利用率</th><th>进度</th><th>目的港</th></tr>')
A('</thead><tbody>')

feeder.sort(key=lambda f:parse_dt(f['pkg'])or datetime.max)
for fi,f in enumerate(feeder,1):
    used=f['bsa_used'];cap=f['bsa_cap'];rem=cap-used
    pct=used/cap*100 if cap else 0
    
    if pct>=90:bar_color='#ef5350';bar_bg='#ffebee'
    elif pct>=60:bar_color='#fb8c00';bar_bg='#fff3e0'
    else:bar_color='#4caf50';bar_bg='#e8f5e9'
    
    parts=[];ds='-'
    for p in ['SOK','JED','MUN','NGB']:
        if f['dst'].get(p):parts.append(p+':'+f['dst'][p])
    ds=', '.join(parts)if parts else '-'
    
    A(f"<tr><td>{fi}</td><td><b>{esc(f['vessel'])}</b></td><td>{esc(f['svc'])}</td><td class='date-f'>{f['pkg']}</td>")
    A(f"<td class='teu-n'>{cap}</td><td class='teu-n'>{used}</td><td class='teu-n'>{rem}</td><td>{pct:.0f}%</td>")
    A(f"<td><div class='progress'style='width:120px'><div class='progress-bar'style='width:{min(pct,100)}%;background:{bar_color}'></div>")
    A(f"<span class='progress-text'>{pct:.0f}%</span></div></td><td>{ds}</td></tr>")

A('</tbody></table></div></div>')

# Footer
A("<div style='text-align:center;color:#aaa;font-size:11px;padding:20px 0'>Generated by Budai Scheduler · Data source: REX & FEEDER SCHEDULE 5.25.xlsx · Sheet: 25_May</div>")
A('</body></html>')

out=r'c:\Users\leahliu\WorkBuddy\20260525175100\budai_plan.html'
with open(out,'w',encoding='utf-8') as f:f.write('\n'.join(L))

print("=" * 50)
print("=== PROFESSIONAL BUDAI PLAN ===")
print(f"Total Cargo:     {total_cargo:,} TEU")
print(f"Assigned:        {total_assigned:,} TEU ({pct_assigned}%)")
print(f"Unassigned:      {total_unassigned:,} TEU ({round(total_unassigned/total_cargo*100)}%)")
print(f"BSA Utilization:{total_bsa_used}/{total_bsa:,} TEU ({pct_bsa}%)")
print(f"Avg Wait Days:   {avg_wait}")
print(f"Assignments:     {len(final_assignments)}")
print(f"Recommendations: {len(recommendations)}")
print(f"\nOutput:          {out}")
