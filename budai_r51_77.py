from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

pkg_col = col_map.get('PKG', 21)
pkge_col = col_map.get('PKG EB', 29)
teu_col = col_map.get('EFF. TEUS', 12)
vessel_col = col_map.get('CUL VESSELS', 3)
week_col = col_map.get('WEEK', 2)
service_col = col_map.get('SERVICES', 5)
voyage_col = col_map.get('VOYAGE .NO', 4)
remarks_col = col_map.get('REMARKS', 34)

dest_cols = {}
for pn in ['SOK','JED','MUN','NGB']:
    if pn in col_map:
        dest_cols[pn] = col_map[pn]

def fmt_date(d):
    if d is None: return ''
    if isinstance(d, datetime): return d.strftime('%Y-%m-%d')
    s = str(d)[:10]
    return s if s and not s.startswith('<') else ''

def get_rgb(cell):
    fill = cell.fill
    try:
        rgb = str(fill.fgColor.rgb) if fill.fgColor and fill.fgColor.rgb else '00000000'
        # openpyxl returns ARGB, normalize
        if len(rgb) == 8:
            return rgb.upper()
        return '00000000'
    except Exception as e:
        return f'ERROR:{e}'

def classify_color(rgb):
    """Classify based on actual RGB value"""
    # Known colors from the file legend
    if rgb == '00000000' or rgb == 'NONE':
        return 'WHITE', '#ffffff', '白底(干线)'
    
    # Pink: FFDAF2D0 - the main pink color from screenshot
    if 'DAF2D0' in rgb:
        return 'PINK', '#DAF2D0', '粉底(布袋)'
    
    # Other pink-ish/orange tones seen in data
    if any(x in rgb for x in ['FBE2D5','CAEDFB']):
        return 'PINK', '#f8d7da', '粉底/橙(接驳)'
    
    # Gray
    if 'BFBFBF' in rgb or 'C0C0C0' in rgb:
        return 'GRAY', '#cccccc', '灰色(TBC)'
        
    return f'OTHER-{rgb}', '#ddd', f'其他({rgb})'

hdr = "R# | RGB           | 分类 | WEEK | VESSEL                  | SVC    | VOY  |   TEU | PKG        | PKGE| SOK | JED | MUN | NGB | REMARKS"
print("=" * len(hdr))
print(hdr)
print("-" * len(hdr))

all_rows = []
for row in range(51, 78):  # R51 to R77 inclusive
    vessel_cell = sheet.cell(row=row, column=vessel_col)
    vessel = str(vessel_cell.value or '').strip()
    if not vessel:
        continue
    
    rgb = get_rgb(vessel_cell)
    ctype, css_color, label = classify_color(rgb)
    
    week = sheet.cell(row=row, column=week_col).value
    svc = sheet.cell(row=row, column=service_col).value
    voyage = sheet.cell(row=row, column=voyage_col).value
    pkg = sheet.cell(row=row, column=pkg_col).value
    pkge = sheet.cell(row=row, column=pkge_col).value
    teu = sheet.cell(row=row, column=teu_col).value
    rmk = sheet.cell(row=row, column=remarks_col).value if remarks_col else None
    
    dest_etas = {}
    for pn, pc in dest_cols.items():
        v = sheet.cell(row=row, column=pc).value
        if v:
            dest_etas[pn] = fmt_date(v)
    
    row_data = {
        'row': row, 'vessel': vessel, 'week': week, 'svc': svc,
        'voyage': voyage, 'pkg': fmt_date(pkg), 'pkge': fmt_date(pkge),
        'teu': teu, 'rgb': rgb, 'type': ctype, 'label': label,
        'css_color': css_color, 'remarks': (str(rmk)[:60] if rmk else ''),
        'dest': dest_etas
    }
    all_rows.append(row_data)
    
    w = str(week)
    s = str(svc)
    v = str(voyage)
    t = str(teu)
    r = (str(rmk)[:50]) if rmk else ''
    print(f"{row:3d} | {rgb:14s} | {label:8s} | {w:>4s} | {vessel:24s} | "
          f"{s:6s} | {v:4s} | {t:>5s} | {fmt_date(pkg):10s} | "
          f"{fmt_date(pkge):4s} | {dest_etas.get('SOK',''):>4s} | {dest_etas.get('JED',''):>4s} | "
          f"{dest_etas.get('MUN',''):>4s} | {dest_etas.get('NGB',''):>4s} | {r}")

print()

# Count
white_count = sum(1 for r in all_rows if r['type'] == 'WHITE')
pink_count = sum(1 for r in all_rows if r['type'] == 'PINK')
other_count = len(all_rows) - white_count - pink_count
print(f"Total rows: {len(all_rows)} | WHITE={white_count} | PINK={pink_count} | OTHER={other_count}")
