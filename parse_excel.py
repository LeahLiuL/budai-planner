# -*- coding: utf-8 -*-
"""
budai_planner Excel数据提取脚本 v2
用openpyxl读取Excel，完整保留单元格颜色和日期信息
分类策略：SERVICE航线(SGX/REX)优先 → 颜色回退 → 手动标记

用法:
  python parse_excel.py "C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx" 25_May_
  python parse_excel.py <excel路径> [sheet名]
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date


def rgb_str(fill):
    """提取填充色RGB字符串"""
    try:
        if not fill or fill.fill_type is None:
            return '00000000'
        fg = fill.fgColor
        if fg and fg.rgb:
            s = str(fg.rgb).upper()
            if len(s) == 8 and s.startswith('FF'):
                s = s[2:]
            elif len(s) == 8 and s.startswith('00'):
                s = s[2:]
            return s
        return '00000000'
    except Exception:
        return '00000000'


def dt_to_str(val):
    """将日期值转为 YYYY-MM-DD 字符串"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    try:
        n = float(val)
        if 25000 < n < 60000:
            base = datetime(1899, 12, 30)
            dt = base + __import__('datetime').timedelta(days=int(n))
            return dt.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass
    s = str(val).strip()
    if len(s) == 10 and '-' in s:
        return s
    return None


def find_header_row(ws):
    """找到CUL VESSELS表头行"""
    for r in range(1, min(ws.max_row + 1, 200)):
        for c in range(1, min(ws.max_column + 1, 50)):
            cell_val = ws.cell(r, c).value
            if cell_val and 'CUL VESSELS' in str(cell_val).upper():
                return r
    return -1


def build_col_map(ws, hdr_row):
    """读取表头列名到列号映射"""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(hdr_row, c).value
        if h:
            col_map[str(h).strip().upper()] = c
    return col_map


def identify_port_cols(ws, hdr_row, pkg_col):
    """识别中国港口列(在PKG前)和红海港口列(在PKG后)
    直接遍历表头行而非col_map，避免重名列被覆盖
    """
    CHN_CODES = {'TXG', 'TAO', 'SHA', 'SHK', 'XMN', 'NGB', 'NAS', 'CKG', 'WHC'}
    RED_CODES = {'SOK', 'JED', 'MUN', 'AQJ', 'KHI', 'PSD', 'ADE', 'HOD', 'DJJ'}

    chn_cols = []
    red_cols = []

    for c in range(1, pkg_col):  # 只遍历PKG之前的列
        h = str(ws.cell(hdr_row, c).value or '').strip().upper()
        clean = ''.join(ch for ch in h if ch.isalpha())
        if clean in CHN_CODES:
            chn_cols.append(c)

    for c in range(pkg_col + 1, ws.max_column + 1):  # PKG之后的列
        h = str(ws.cell(hdr_row, c).value or '').strip().upper()
        clean = ''.join(ch for ch in h if ch.isalpha())
        if clean in RED_CODES:
            red_cols.append(c)
        elif any(kw in h for kw in ['KHI', 'AQJ']):
            red_cols.append(c)

    return chn_cols, red_cols


def has_date_in_cols(ws, row, cols):
    """检查该行在指定列中是否有日期值"""
    for c in cols:
        v = ws.cell(row, c).value
        if v and isinstance(v, (datetime, date)):
            return True
    return False


def classify_ship(svc, color, has_chn, has_red):
    """
    分类逻辑：SERVICE航线优先 → 颜色回退 → 未知
    返回 (type, typeLabel, fromPort, toPort, classifiedBy)
    type: W=干线 F=布袋 P=直航 ?=未知
    """
    if 'SGX' in svc:
        return 'W', '干线(SGX)', '中国各港', '', 'SVC:SGX→干线'

    if 'REX' in svc:
        if has_chn and has_red:
            return 'P', '直航(REX中+红)', '中国各港', '红海(直达)', 'SVC:REX→直航(中+红)'
        elif not has_chn and has_red:
            return 'F', '布袋(REX仅红)', '巴生PKG', '红海', 'SVC:REX→布袋(仅红)'
        elif has_chn and not has_red:
            return 'W', '干线(REX仅中)', '中国各港', '', 'SVC:REX→干线(仅中)'
        else:
            return 'W', '干线(REX无日期)', '中国各港', '', 'SVC:REX→干线(无日期)'

    if svc:
        # 其他航线(AEM/RES/CGX等)，尝试颜色
        if 'DAF2D0' in color:
            return 'P', '直航(颜色绿)', '中国各港', '红海(直达)', f'color:DAF2D0→直航 [svc:{svc}]'
        elif 'FBE2D5' in color:
            return 'F', '布袋(颜色橙)', '巴生PKG', '红海', f'color:FBE2D5→布袋 [svc:{svc}]'
        elif not color or color == '00000000' or color == 'FFFFFFFF':
            return 'W', '干线(颜色白)', '中国各港', '', f'color:白→干线 [svc:{svc}]'
        else:
            return '?', f'未知(svc:{svc},color:{color})', '—', '—', f'unknown [svc:{svc}]'

    # 无SERVICE → 纯颜色分类
    if 'DAF2D0' in color:
        return 'P', '直航', '—', '红海', 'color:DAF2D0→直航'
    elif 'FBE2D5' in color:
        return 'F', '布袋', '巴生PKG', '红海', 'color:FBE2D5→布袋'
    elif not color or color == '00000000' or color == 'FFFFFFFF':
        return 'W', '干线', '—', '', 'color:白→干线'
    else:
        return '?', f'未知({color})', '—', '—', f'unknown color:{color}'


def parse_excel(filepath, sheet_name=None):
    wb = load_workbook(filepath, data_only=True)

    # 自动选sheet: 优先选有CUL VESSELS的最新sheet（排除review/option/trade drafts）
    if sheet_name:
        ws = wb[sheet_name]
    else:
        candidates = []
        for name in wb.sheetnames:
            ws = wb[name]
            hdr = find_header_row(ws)
            if hdr > 0:
                nl = name.lower()
                # 排除draft/review/option sheet
                is_draft = any(kw in nl for kw in ['review', 'option', 'draft', 'tentative', 'sheet'])
                # 月份评分: May > Apr > Mar
                month_score = 0
                if 'may' in nl: month_score = 5
                elif 'apr' in nl: month_score = 4
                elif 'mar' in nl: month_score = 3
                elif 'jun' in nl: month_score = 2
                candidates.append((name, month_score, is_draft, ws.max_row, ws.max_column))
        if not candidates:
            print("ERROR: 所有Sheet中都找不到CUL VESSELS表头")
            sys.exit(1)
        # Sort: not-draft first, then month desc, then rows desc
        candidates.sort(key=lambda x: (x[2], -x[1], -x[3]))
        sheet_name = candidates[0][0]
        ws = wb[sheet_name]

    hdr_row = find_header_row(ws)
    if hdr_row < 0:
        print("ERROR: 找不到表头(CUL VESSELS列)")
        sys.exit(1)

    col_map = build_col_map(ws, hdr_row)
    print(f"Sheet: {sheet_name} | Header row: {hdr_row} | {ws.max_row} rows x {ws.max_column} cols")

    # 关键列
    ves_col = col_map.get('CUL VESSELS', col_map.get('VESSEL', 0))
    svc_col = col_map.get('SERVICES', col_map.get('SVC', 0))
    pkg_col = col_map.get('PKG', 0)
    teu_col = col_map.get('EFF. TEUS', col_map.get('EFF.TEUS', col_map.get('TEU', 0)))
    pol_col = col_map.get('POL', 0)
    pod_col = col_map.get('POD', 0)
    disch_vol_col = col_map.get('DISCHARGE VOL', 0)

    if not ves_col:
        print("ERROR: 找不到CUL VESSELS列")
        sys.exit(1)

    # 港口列识别
    chn_port_cols, red_port_cols = identify_port_cols(ws, hdr_row, pkg_col)
    print(f"中国港口列: {chn_port_cols} | 红海港口列: {red_port_cols}")

    result = {
        'whites': [], 'feeders': [], 'pinks': [],
        'meta': {
            'source': filepath,
            'sheet': sheet_name,
            'headerRow': hdr_row,
            'columns': {k: v for k, v in col_map.items()
                        if k in ['CUL VESSELS', 'PKG', 'PKGE', 'EFF. TEUS',
                                 'SERVICES', 'POL', 'POD', 'DISCHARGE VOL']}
        }
    }

    stats = {'SVC': 0, 'color': 0, 'unknown': 0}

    for r in range(hdr_row + 1, ws.max_row + 1):
        ves = ws.cell(r, ves_col).value
        if not ves or not str(ves).strip():
            continue

        name = str(ves).strip()
        color = rgb_str(ws.cell(r, ves_col).fill)

        pkg_date = dt_to_str(ws.cell(r, pkg_col).value) if pkg_col else None

        teu_val = ws.cell(r, teu_col).value if teu_col else 0
        try:
            teu_num = int(float(teu_val)) if teu_val else 0
        except (ValueError, TypeError):
            teu_num = 0

        disch_val = ws.cell(r, disch_vol_col).value if disch_vol_col else None
        try:
            disch_num = int(float(disch_val)) if disch_val else 0
        except (ValueError, TypeError):
            disch_num = 0

        svc = str(ws.cell(r, svc_col).value).strip().upper() if svc_col else ''
        pol = str(ws.cell(r, pol_col).value).strip() if pol_col else ''
        pod = str(ws.cell(r, pod_col).value).strip() if pod_col else ''

        has_chn = has_date_in_cols(ws, r, chn_port_cols)
        has_red = has_date_in_cols(ws, r, red_port_cols)

        ship_type, type_label, from_port, to_port, classified_by = classify_ship(
            svc, color, has_chn, has_red
        )

        # 统计
        if classified_by.startswith('SVC:'):
            stats['SVC'] += 1
        elif classified_by.startswith('color:'):
            stats['color'] += 1
        else:
            stats['unknown'] += 1

        entry = {
            'row': r,
            'name': name,
            'type': ship_type,
            'typeLabel': type_label,
            'pkg': pkg_date,
            'teu': teu_num,
            'dischargeVol': disch_num,
            'svc': svc,
            'color': color,
            'fromPort': from_port or pol or '中国各港',
            'toPort': to_port or pod or '',
            'pol': pol or '—',
            'pod': pod or '—',
            'classifiedBy': classified_by
        }

        if ship_type == 'W':
            result['whites'].append(entry)
        elif ship_type == 'F':
            result['feeders'].append(entry)
        elif ship_type == 'P':
            result['pinks'].append(entry)
        else:
            result['whites'].append(entry)  # 未知先放入whites供手动分类

    result['summary'] = {
        'whites': len(result['whites']),
        'feeders': len(result['feeders']),
        'pinks': len(result['pinks']),
        'total': len(result['whites']) + len(result['feeders']) + len(result['pinks']),
        'classifiedBy': stats
    }

    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python parse_excel.py <xlsx文件> [sheet名]")
        print("例: python parse_excel.py schedule.xlsx 25_May_")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None

    data = parse_excel(xlsx_path, sheet)

    out_path = xlsx_path.rsplit('.', 1)[0] + '_parsed.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n=== 解析完成 ===")
    s = data['summary']
    stats = s.get('classifiedBy', {})
    print(f"干线条: {s['whites']} 艘")
    print(f"布袋船: {s['feeders']} 艘")
    print(f"直航船: {s['pinks']} 艘")
    print(f"总计:   {s['total']} 艘")
    print(f"分类依据: 航线规则={stats.get('SVC',0)} 颜色={stats.get('color',0)} 未知={stats.get('unknown',0)}")
    print(f"\nJSON已保存: {out_path}")
