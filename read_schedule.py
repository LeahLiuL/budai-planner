from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

# Header is in Row 7
header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

print('=== Column Map ===')
for k, v in sorted(col_map.items()):
    print(f'  {k}: Col {get_column_letter(v)}')
print()

# Key column indices
pkg_col = col_map.get('PKG', None)           # Col U - domestic to PKG ETA
pkge_col = col_map.get('PKG EB', None)       # Col AC - EB to PKG ETA  
teu_col = col_map.get('EFF. TEUS', None)     # Col L - capacity
vol_col = col_map.get('DISCHARGE VOL', None)  # Col V - WB cargo volume
vessel_col = col_map.get('CUL VESSELS', None) # Col C
week_col = col_map.get('WEEK', None)          # Col B
service_col = col_map.get('SERVICES', None)   # Col E

print(f'Key cols: PKG={pkg_col}, PKGEB={pkge_col}, TEU={teu_col}, VOL={vol_col}, VESSEL={vessel_col}')
print()

# Parse all data rows
print('=' * 120)
print(f'{"R#":>3} | {"COLOR":18s} | {"WEEK":4s} | {"VESSEL":24s} | {"SVC":5s} | {"ETA_PKG":10s} | {"ETA_PKGE":10s} | TEU | VOL')
print('-' * 120)

white_rows = []   # main line (domestic -> PKG)
pink_rows = []    # feeder (PKG -> destination)

for row in range(header_row + 1, sheet.max_row + 1):
    vessel_cell = sheet.cell(row=row, column=vessel_col or 3)
    vessel = str(vessel_cell.value or '').strip()
    
    if not vessel:
        continue
    
    # Get background color
    fill = vessel_cell.fill
    rgb = str(fill.fgColor.rgb) if fill.fgColor and fill.fgColor.rgb else 'NONE'
    
    # Determine type by color
    if rgb == 'FFDAF2D0' or 'FF' + (rgb[2:] if len(str(rgb)) >= 6 else '') == 'FFDAF2D0':
        row_type = 'FEEDER'
        is_pink = True
    elif rgb == '00000000' or rgb == 'None' or rgb == 'NONE':
        row_type = 'MAINLINE'
        is_pink = False
    else:
        # Check for pink-like colors
        is_pink = 'DAF2' in str(rgb).upper() or 'F2D0' in str(rgb).upper() or 'FFC0' in str(rgb).upper() or 'FFFFCB' in str(rgb).upper() or 'C000' in str(rgb).upper()
        row_type = 'FEEDER' if is_pink else 'MAINLINE'
    
    week = sheet.cell(row=row, column=week_col or 2).value
    svc = sheet.cell(row=row, column=service_col or 5).value
    pkg_eta = sheet.cell(row=row, column=pkg_col).value if pkg_col else None
    pkge_eta = sheet.cell(row=row, column=pkge_col).value if pkge_col else None
    teu = sheet.cell(row=row, column=teu_col).value if teu_col else None
    vol = sheet.cell(row=row, column=vol_col).value if vol_col else None
    
    # Format dates
    pkg_str = pkg_eta.strftime('%d-%b') if isinstance(pkg_eta, datetime) else str(pkg_eta)[:8] if pkg_eta else ''
    pkge_str = pkge_eta.strftime('%d-%b') if isinstance(pkge_eta, datetime) else str(pkge_eta)[:8] if pkge_eta else ''
    
    print(f'{row:3d} | {rgb:18s} | {str(week):4s} | {vessel:24s} | {str(svc):5s} | {pkg_str:10s} | {pkge_str:10s} | {str(teu):4s} | {str(vol):4s}')
    
    row_data = {
        'row': row,
        'vessel': vessel,
        'week': week,
        'service': svc,
        'eta_pkg': pkg_eta,
        'eta_pkge': pkge_eta,
        'teu': teu,
        'volume': vol,
        'color_rgb': rgb,
        'type': row_type
    }
    
    if is_pink:
        pink_rows.append(row_data)
    else:
        white_rows.append(row_data)

print()
print(f'SUMMARY: {len(white_rows)} MAINLINE (white), {len(pink_rows)} FEEDER (pink)')
print()

# Show pink rows details
if pink_rows:
    print('=== PINK ROWS (Feeder/接驳船) Details ===')
    for r in pink_rows:
        print(f"  R{r['row']:02d}: {r['vessel']:22s} | SVC={r['service']} | PKG={r['eta_pkg']} | TEU={r['teu']} | VOL={r['volume']}")
