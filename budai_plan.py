from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7

col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

pkg_col = col_map.get('PKG', 21)
pkge_col = col_map.get('PKG EB', 29)
teu_col = col_map.get('EFF. TEUS', 12)
vessel_col = col_map.get('CUL VESSELS', 3)
week_col = col_map.get('WEEK', 2)
service_col = col_map.get('SERVICES', 5)
voyage_col = col_map.get('VOYAGE .NO', 4)
remarks_col = col_map.get('REMARKS', 34)

dest_cols = {}
for pn in ['SOK', 'JED', 'MUN', 'NGB']:
    if pn in col_map:
        dest_cols[pn] = col_map[pn]

def fmt_date(d):
    if d is None: return ''
    if isinstance(d, datetime): return d.strftime('%Y-%m-%d')
    s = str(d)[:10]
    try:
        dt = datetime.strptime(s, '%Y-%m-%d')
        return dt.strftime('%Y-%m-%d')
    except: return s[:10] if s else ''

def parse_dt(s):
    if not s or str(s).strip() == '' or str(s).strip() == 'X': return None
    try: return datetime.strptime(str(s)[:10], '%Y-%m-%d')
    except:
        try: return datetime.strptime(str(s), '%Y-%m-%d %H:%M:%S')
        except: return None

# Find ZHI YING HE SHUN target row (PKG >= Jun 7)
target_row = None
all_rows_data = []

for row in range(header_row + 1, sheet.max_row + 1):
    vessel_cell = sheet.cell(row=row, column=vessel_col)
    vessel = str(vessel_cell.value or '').strip()
    if not vessel: continue
    
    fill = vessel_cell.fill
    rgb_raw = ''
    try:
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb_raw = str(fill.fgColor.rgb)
    except:
        rgb_raw = '00000000'
    
    pkg_val = fmt_date(sheet.cell(row=row, column=pkg_col).value if pkg_col else None)
    
    # Detect target: ZHI YING HE SHUN with PKG ~ Jun 7
    if 'ZHI YING HE SHUN' in vessel.upper() or ('ZYHS' in vessel.upper() and 'CHONGFU' not in vessel.upper()):
        if pkg_val and pkg_val >= '2026-06-05' and pkg_val <= '2026-06-10':
            target_row = row
    
    week = sheet.cell(row=row, column=week_col).value
    svc = sheet.cell(row=row, column=service_col).value
    voyage = sheet.cell(row=row, column=voyage_col).value
    pkge_val = fmt_date(sheet.cell(row=row, column=pkge_col).value if pkge_col else None)
    teu = sheet.cell(row=row, column=teu_col).value if teu_col else None
    remarks = sheet.cell(row=row, column=remarks_col).value if remarks_col else None
    
    dest_etas = {}
    for pname, pc in dest_cols.items():
        v = sheet.cell(row=row, column=pc).value
        if v: dest_etas[pname] = fmt_date(v)
    
    all_rows_data.append({
        'row': row,
        'vessel': vessel,
        'week': str(week) if week else '',
        'svc': str(svc) if svc else '',
        'voyage': str(voyage) if voyage else '',
        'eta_pkg': pkg_val,
        'eta_pkge': pkge_val,
        'teu': teu,
        'color_rgb': rgb_raw,
        'remarks': str(remarks)[:100] if remarks else '',
        'dests': dest_etas
    })

print(f"Target ZYHS found at Row: {target_row}")
if target_row:
    print(f"  Vessel: {all_rows_data[target_row-header_row]['vessel']}, PKG: {all_rows_data[target_row-header_row]['eta_pkg']}")

# Get all rows after target_row
start_idx = None
for i, r in enumerate(all_rows_data):
    if r['row'] == target_row:
        start_idx = i
        break

after_rows = all_rows_data[start_idx:] if start_idx is not None else all_rows_data

print(f"\nRows after target: {len(after_rows)}")

# Classify by color - show actual RGB for verification
def classify_color(rgb):
    if not rgb or rgb == '00000000' or rgb == '':
        return ('WHITE', '#ffffff', '白底=国内→巴生')
    rgb_upper = rgb.upper().replace('FF','') if len(rgb) == 8 else rgb.upper()
    
    # Known pink colors from this spreadsheet
    pink_codes = ['DAF2D0', 'FBE2D5', 'CAEDFB']
    for p in pink_codes:
        if p in rgb_upper:
            return ('PINK', '#DAF2D0', '粉底=巴生→红海')
    
    # Check if it starts with FF (non-black)
    if len(rgb) == 8 and rgb[:2] == 'FF' and rgb[2:] != '000000':
        return ('PINK', '#DAF2D0', '粉底=巴生→红海')
    
    return ('UNKNOWN', '#cccccc', '未知')

# Build HTML
html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>布袋船计划 - ZYHS之后全部船期</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:#f0f2f5;padding:16px;color:#333;font-size:13px}
h1{font-size:18px;color:#1565c0;margin-bottom:4px}
.sub{font-size:12px;color:#666;margin-bottom:14px}
.card{background:#fff;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.12);margin-bottom:14px;overflow:hidden}
.hd{padding:10px 14px;font-weight:bold;font-size:14px;display:flex;align-items:center;gap:8px;border-bottom:1px solid #eee}
.badge{display:inline-block;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:normal}
.b-white{background:#e8eaed;color:#333}.b-pink{background:#fce4ec;color:#c2185b}
table{width:100%;border-collapse:collapse;font-size:11px}
th{background:#f5f5f5;padding:6px 8px;text-align:left;font-weight:600;border-bottom:2px solid #ddd;position:sticky;top:0}
td{padding:5px 8px;border-bottom:1px solid #f0f0f0;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
tr:hover td{background:#f8fbff}
.date{font-family:"Courier New",monospace;font-size:10.5px}
.teu{text-align:right;font-weight:600}
.color-swatch{display:inline-block;width:20px;height:16px;border-radius:3px;border:1px solid #ccc;vertical-align:middle;margin-right:4px}
.rgb-tag{font-size:9px;color:#999;font-family:monospace;background:#f5f5f5;padding:1px 4px;border-radius:2px}
.remark{color:#888;font-style:italic;font-size:10px;max-width:250px}
.w-ok{color:#2e7d32;font-weight:700}.w-warn{color:#e65100;font-weight:700}.w-bad{color:#c62828;font-weight:700}
.match-card{border:1px solid #ddd;border-radius:6px;margin:8px 14px;overflow:hidden}
.m-hd{background:#e3f2fd;padding:7px 10px;font-size:12px;font-weight:bold;display:flex;align-items:center;justify-content:space-between}
.m-body{padding:6px 10px}
.m-row{display:flex;align-items:center;padding:4px 0;font-size:11px;border-bottom:1px dotted #e0e0e0;gap:6px}
.m-row:last-child{border:none}
.plan-box{margin:14px}
.plan-item{background:#fffbe6;border-left:4px solid #ffc107;padding:10px 14px;margin:8px 0;border-radius:0 6px 6px 0;font-size:12px;line-height:1.6}
.plan-best{border-left-color:#4caf50;background:#e8f5e9}
.legend{padding:10px 14px;font-size:11.5px;display:flex;gap:24px;border-top:1px solid #eee}
.l-itm{display:flex;align-items:center;gap:5px}
</style>
</head>
<body>

<div class="card">
<div class="hd">📋 布袋船衔接计划 — <span style="font-weight:normal;font-size:12px">ZYHS 之后全部船期</span></div>
<div class="legend">
<span class="l-itm"><span class="color-swatch" style="background:#fff"></span><b>白色底</b> = 干线船（国内→巴生，送货到PKG）</span>
<span class="l-itm"><span class="color-swatch" style="background:#DAF2D0"></span><b>粉色底</b> = 布袋船（巴生→红海，从PKG接走）</span>
<span class="l-itm" style="margin-left:auto;color:#666">操作窗口：≥1天 | 快转：≤2天</span>
</div>
'''

# Section 1: All rows after ZYHS with color verification
white_ships = []
pink_ships = []

html += '<div style="padding:14px"><h3 style="font-size:13px;margin-bottom:8px">📊 全部船期列表（含颜色RGB验证）</h3>'
html += '<table><thead><tr><th>R#</th><th>颜色</th><th>RGB</th><th>分类</th><th>VESSEL</th><th>SVC</th><th>TEU</th>'
html += '<th>ETA_PKG</th><th>SOK</th><th>JED</th><th>MUN</th><th>NGB</th><th>REMARKS</th></tr></thead><tbody>\n'

for r in after_rows:
    ctype, ccolor, clabel = classify_color(r['color_rgb'])
    
    html += f"<tr>"
    html += f"<td>{r['row']}</td>"
    html += f"<td><span class='color-swatch' style='background:{ccolor}'></span></td>"
    html += f"<td class='rgb-tag'>{r['color_rgb']}</td>"
    html += f"<td>{clabel}</td>"
    html += f"<td><b>{r['vessel']}</b></td>"
    html += f"<td>{r['svc']}</td>"
    html += f"<td class='teu'>{r['teu']}</td>"
    html += f"<td class='date'>{r['eta_pkg']}</td>"
    for pn in ['SOK','JED','MUN','NGB']:
        html += f"<td class='date'>{r['dests'].get(pn,'')}</td>"
    html += f"<td class='remark'>{r['remarks']}</td>"
    html += f"</tr>\n"
    
    if ctype == 'WHITE' and r['eta_pkg']:
        white_ships.append(r)
    elif ctype == 'PINK' and r['eta_pkg']:
        pink_ships.append(r)

html += '</tbody></table></div></div>\n'

# Section 2: Connection Plan
html += f'<div class="card"><div class="hd">🔗 布袋衔接计划（白底到PKG → 粉底离开PKG）</div>'

html += f'<div style="padding:8px 14px;font-size:11.5px;color:#555;">'
html += f'干线船（白底）<b>{len(white_ships)}</b> 艘 | 布袋船（粉底）<b>{len(pink_ships)}</b> 艘'
html += '</div>'

total_connections = 0
plan_items = []

for ws in white_ships:
    ws_eta = parse_dt(ws['eta_pkg'])
    if not ws_eta:
        continue
    
    matches = []
    for ps in pink_ships:
        ps_etd = parse_dt(ps['eta_pkg'])
        if not ps_etd:
            continue
        
        delta = (ps_etd - ws_eta).days
        if delta >= 1:  # minimum operation window
            wcls = 'w-ok' if delta <= 2 else ('w-warn' if delta <= 4 else 'w-bad')
            matches.append({'feeder': ps, 'delta': delta, 'cls': wcls})
    
    if matches:
        matches.sort(key=lambda x: x['delta'])
        best = matches[0]
        total_connections += 1
        
        plan_items.append({
            'main': ws,
            'best_feeder': best['feeder'],
            'best_delta': best['delta'],
            'best_cls': best['cls'],
            'all_matches': matches
        })
        
        html += '<div class="match-card">'
        html += f"<div class='m-hd'>"
        html += f"<span><span class='color-swatch' style='background:#fff'></span> <b>{ws['vessel']}</b> ({ws['svc']})</span>"
        html += f"<span class='date'>到PKG: {ws['eta_pkg']}</span> <span class='{best['cls']}'>最佳等 {best['delta']}天</span>"
        html += "</div>"
        
        html += "<div class='m-body'>"
        for m in matches[:10]:  # top 10 feeders per mainline
            f = m['feeder']
            dest_str = ', '.join([f"{k}:{v}" for k,v in f['dests'].items() if v])[:60]
            html += f"<div class='m-row'>"
            html += f"<span class='color-swatch' style='background:#DAF2D0;width:14px;height:12px;'></span>"
            html += f"<span style='width:150px'><b>{f['vessel']}</b> ({f['svc']})</span>"
            html += f"<span class='date' style='width:75px'>ETD {f['eta_pkg']}</span>"
            wclass = m["cls"]
            html += f"<span class='{wclass}' style='width:50px;text-align:center'>等{m['delta']}天</span>"
            html += f"<span style='width:45px;text-align:right'>TEU:{f['teu']}</span>"
            html += f"<span style='color:#777;font-size:10px'>{dest_str}</span>"
            html += f"</div>"
        if len(matches) > 10:
            html += f"<div style='font-size:10px;color:#aaa;padding:2px 0;'>+ 还有 {len(matches)-10} 条可衔接</div>"
        html += "</div></div>\n"

# Summary plan box
html += f'<div class="plan-box">'
html += f'<h3 style="font-size:13px;margin:8px 0 10px">📌 排计划建议汇总</h3>'

# Sort plan items by best delta (shortest first)
plan_items.sort(key=lambda x: x['best_delta'])

for i, pi in enumerate(plan_items):
    m = pi['main']
    b = pi['best_feeder']
    d = pi['best_delta']
    cls = pi['best_cls']
    
    best_tag = '★ 最佳' if d <= 2 else ('✓ 可用' if d <= 4 else '△ 较长')
    
    html += f'<div class="plan-item {"plan-best" if d <= 2 else ""}">'
    html += f'<b>{i+1}.</b> <b>{m["vessel"]}</b> 到PKG <code>{m["eta_pkg"]}</code> → '
    html += f'接 <b>{b["vessel"]}</b> ETD <code>{b["eta_pkg"]}</code> '
    html += f'| 等待 <span class="{cls}">{d}天</span> {best_tag}'
    
    # show destination
    if b['dests']:
        dest_info = ' | 目的港: ' + ', '.join([f"{k}({v[5:]})" for k,v in b['dests'].items()])
        html += dest_info
    
    if b['teu']:
        html += f' | 舱位 {b["teu"]}TEU'
    
    html += '</div>\n'

if not plan_items:
    html += '<p style="padding:14px;color:#c62828;">⚠️ 未找到有效衔接，请检查颜色分类和日期数据</p>'

html += '</div>'  # plan_box
html += '</div>'  # card

html += '</body></html>'

outpath = r'c:\Users\leahliu\WorkBuddy\20260525175100\budai_plan.html'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\nDone! Output: {outpath}")
print(f"White ships (domestic->PKG): {len(white_ships)}")
print(f"Pink ships (PKG->RedSea): {len(pink_ships)}")
print(f"Total connections found: {total_connections}")
