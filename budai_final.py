from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook(r'C:/CULINES/Claw Report/REX & FEEDER SCHEDULE 5.25.xlsx', data_only=True)
sheet = wb['25_May_']

header_row = 7
col_map = {}
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=header_row, column=c).value
    if h:
        col_map[h.strip().upper()] = c

pkg_col = col_map.get('PKG', 21)
pkge_col = col_map.get('PKG EB', 29)
teu_col = col_map.get('EFF. TEUS', 12)
vessel_col = col_map.get('CUL VESSELS', 3)
week_col = col_map.get('WEEK', 2)
service_col = col_map.get('SERVICES', 5)
voyage_col = col_map.get('VOYAGE .No', 4)
remarks_col = col_map.get('REMARKS', 34)

dest_cols = {}
for pn in ['SOK','JED','MUN','NGB']:
    if pn in col_map:
        dest_cols[pn] = col_map[pn]

def fmt(d):
    if d is None: return ''
    if isinstance(d,datetime): return d.strftime('%Y-%m-%d')
    s=str(d)[:10]
    return s if s else ''

def get_rgb(cell):
    try:
        r=str(cell.fill.fgColor.rgb)
        return r.upper() if len(r)==8 else '00000000'
    except: return '00000000'

def classify(rgb):
    if rgb=='00000000': return ('WHITE','#ffffff','白底(国内->巴生)')
    if 'DAF2D0' in rgb: return ('PINK','#DAF2D0','粉底(国内->红海直航)')
    if any(x in rgb for x in ['FBE2D5']): return ('FEEDER','#FBE2D5','橙底(布袋船/巴生->红海)')
    return ('OTHER','#ddd','其他')

# Collect R51-R77
all_rows = []
for row in range(51,78):
    vc=sheet.cell(row=row,column=vessel_col)
    v=str(vc.value or '').strip()
    if not v: continue
    
    rgb=get_rgb(vc); ctype,css,lab=classify(rgb)
    
    pkg=sheet.cell(row=row,column=pkg_col).value
    pkge=sheet.cell(row=row,column=pkge_col).value
    teu=sheet.cell(row=row,column=teu_col).value
    wk=sheet.cell(row=row,column=week_col).value
    svc=sheet.cell(row=row,column=service_col).value
    voy=sheet.cell(row=row,column=voyage_col).value
    rmk=sheet.cell(row=row,column=remarks_col).value
    
    dst={}
    for pn,pc in dest_cols.items():
        val=sheet.cell(row=row,column=pc).value
        if val: dst[pn]=fmt(val)
    
    all_rows.append({
        'row':row,'vessel':v,'week':wk,'svc':svc,'voy':voy,
        'pkg':fmt(pkg),'pkge':fmt(pkge),'teu':teu,
        'rgb':rgb,'ctype':ctype,'label':lab,'color':css,
        'rmk':(str(rmk)[:60] if rmk else ''),'dst':dst
    })

white=[r for r in all_rows if r['ctype']=='WHITE']
feeders=[r for r in all_rows if r['ctype']=='FEEDER']  # only orange = 布袋船
pink=[r for r in all_rows if r['ctype']=='PINK']  # 直航船，不参与衔接

# Connection matching
def parse_dt(s):
    if not s: return None
    try: return datetime.strptime(s,'%Y-%m-%d')
    except: return None

connections=[]
for mw in white:
    eta=parse_dt(mw['pkg'])
    if not eta: continue
    matches=[]
    for mp in feeders:
        etd=parse_dt(mp['pkg'])
        if etd and (etd-eta).days>=1:
            delta=(etd-eta).days
            cls='good' if delta<=3 else ('warn' if delta<=5 else 'bad')
            matches.append({'f':mp,'delta':delta,'cls':cls})
    matches.sort(key=lambda x:x['delta'])
    connections.append({'main':mw,'matches':matches})

connections.sort(key=lambda c: parse_dt(c['main']['pkg']) or datetime.max)

# Build HTML - use string concat instead of complex f-strings
lines = []
L = lines.append

L('<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8">')
L('<title>布袋船计划 - R51~R77</title>')
L('''<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:#f5f6fa;padding:20px;color:#333}
.card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);margin-bottom:20px;overflow:hidden}
.card-head{padding:14px 18px;font-weight:bold;font-size:16px;display:flex;align-items:center;gap:10px;border-bottom:1px solid #eee}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#fafbfc;padding:8px 10px;text-align:left;font-weight:600;border-bottom:2px solid #e1e4e8}
td{padding:6px 10px;border-bottom:1px solid #f0f0f0;white-space:nowrap}
tr:hover td{background:#f8f9ff}
.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:normal}
.bg-w{background:#e8eaed;color:#444} .bg-p{background:#fce4ec;color:#c2185b} .bg-f{background:#fff3e0;color:#e65100}
.date-f{font-family:"Courier New",monospace;font-size:12px}
.teu-n{font-weight:600;text-align:right}
.good{color:#2e7d32;font-weight:600} .warn{color:#e65100;font-weight:600} .bad{color:#c62828;font-weight:600}
.rmks{color:#888;font-style:italic;font-size:11px;max-width:250px;overflow:hidden;text-overflow:ellipsis}
.match-card{border:1px solid #e0e0e0;border-radius:8px;margin-bottom:14px;overflow:hidden}
.match-head{background:#e3f2fd;padding:10px 14px;display:flex;align-items:center;gap:12px;font-weight:bold;font-size:14px}
.match-body{padding:10px 14px}
.m-row{display:flex;align-items:center;padding:5px 0;border-bottom:1px dotted #e8e8e8;font-size:12px}
.m-row:last-child{border:none}
.summary{display:flex;gap:24px;padding:18px;flex-wrap:wrap}
.s-item .n{font-size:28px;font-weight:bold;color:#1565c0}
.s-item .l{font-size:12px;color:#666}
.color-box{width:22px;height:22px;border-radius:4px;border:1px solid #ccc;display:inline-block;vertical-align:middle}
.legend-bar{padding:10px 18px;display:flex;gap:24px;font-size:12px;border-top:1px solid #eee;background:#fafafa}
</style></head><body>
''')

# Summary header
L('<div class="card">')
L('<div class="card-head">📋 布袋船排计划 — R51~R77（ZHI YING HE SHUN 6月7日之后）</div>')
L('<div class="summary">')
L('<div class="s-item"><div class="n">' + str(len(white)) + '</div><div class="l">白底 · 国内→巴生</div></div>')
L('<div class="s-item"><div class="n">' + str(len(feeders)) + '</div><div class="l">橙底 · 布袋船(巴生→红海)</div></div>')
L('<div class="s-item"><div class="n">' + str(len(pink)) + '</div><div class="l">粉底 · 国内→红海直航</div></div>')
L('</div>')
L('<div class="legend-bar">')
L('<span>🟦 <span class="color-box" style="background:#fff"></span> 白底 = 干线船（国内港口→巴生卸货）</span>')
L('<span>🟩 <span class="color-box" style="background:#DAF2D0"></span> 粉底 = 直航船（国内→红海，不经巴生）</span>')
L('<span>🟧 <span class="color-box" style="background:#FBE2D5"></span> 橙底 = 布袋船（巴生装货→红海）</span>')
L('<span style="margin-left:auto;color:#666">衔接关系：白底(到PKG) → 橙底布袋船(离PKG) | 操作窗口≥1天</span>')
L('</div></div>')

# Section 1: Full table
L('<div class="card"><div class="card-head">📊 全部船期明细（R51~R77，含颜色分类验证）</div>')
L('<div style="overflow-x:auto"><table><thead><tr><th>R#</th><th>颜色RGB</th><th>分类</th><th>VESSEL</th><th>SVC</th><th>TEU</th><th>PKG</th><th>PKGE</th><th>SOK</th><th>JED</th><th>MUN</th><th>NGB</th><th>REMARKS</th></tr></thead><tbody>')

for r in all_rows:
    d=r['dst']
    bg = 'w' if r['ctype']=='WHITE' else ('f' if r['ctype']=='FEEDER' else ('p' if r['ctype']=='PINK' else 'o'))
    L('<tr><td>' + str(r['row']) + '</td><td style="font-size:10px;font-family:monospace">' + r['rgb'] + '</td>')
    L('<td><span class="badge bg-' + bg + '">' + r['label'] + '</span></td>')
    L('<td><strong>' + r['vessel'] + '</strong></td><td>' + str(r['svc']) + '</td><td class="teu-n">' + str(r['teu']) + '</td>')
    L('<td class="date-f">' + r['pkg'] + '</td><td class="date-f">' + r['pkge'] + '</td>')
    L('<td class="date-f">' + d.get('SOK','') + '</td><td class="date-f">' + d.get('JED','') + '</td>')
    L('<td class="date-f">' + d.get('MUN','') + '</td><td class="date-f">' + d.get('NGB','') + '</td>')
    L('<td class="rmks">' + r['rmk'] + '</td></tr>')

L('</tbody></table></div></div>')

# Section 2: Connection plan
L('<div class="card"><div class="card-head">🔗 衔接计划 — 干线船到PKG后最优布袋船匹配</div><div style="padding:16px">')
L('<p style="font-size:13px;color:#555;margin-bottom:14px">每条白底干线船到达巴生后，按<b>等待时间最短</b>排序推荐可衔接的粉/橙底布袋船。</p>')

total_conn = 0
for c in connections:
    m=c['main']; ms=c['matches']
    total_conn += len(ms)
    
    L('<div class="match-card">')
    L('<div class="match-head">')
    L('<span style="background:#e8f5e9;padding:4px 10px;border-radius:4px;">' + m['vessel'] + '</span>')
    L('<small>ETA PKG: <code>' + m['pkg'] + '</code></small>')
    L('<small>| ' + str(m['svc']) + ' | TEU:' + str(m['teu']) + '</small>')
    if ms:
        best=ms[0]
        L('<span style="margin-left:auto" class="' + best['cls'] + '">最佳等 ' + str(best['delta']) + '天</span>')
    else:
        L('<span style="margin-left:auto;color:#999">⚠️ 无可衔接船</span>')
    L('</div>')  # match-head
    
    if ms:
        L("<div class='match-body'>")
        for mm in ms[:10]:
            ff=mm['f']; dd=ff['dst']
            parts=[]
            for p in ['SOK','JED','MUN','NGB']:
                if dd.get(p): parts.append(p+':'+dd[p])
            dest_s = ', '.join(parts) if parts else '-'
            
            L("<div class='m-row'>")
            L('<span style="width:200px"><b>' + ff['vessel'] + '</b> <small>(' + str(ff['svc']) + ')</small></span>')
            L('<span class="date-f" style="width:90px">ETD PKG ' + ff['pkg'] + '</span>')
            L('<span class="' + mm['cls'] + '" style="width:60px;text-align:center">等' + str(mm['delta']) + '天</span>')
            L('<span style="width:60px;text-align:right">TEU:' + str(ff['teu']) + '</span>')
            L('<span style="color:#666;margin-left:8px">→ ' + dest_s + '</span>')
            L("</div>")  # m-row
        if len(ms)>10:
            L('<div style="font-size:11px;color:#999;padding:4px 0">... 还有 ' + str(len(ms)-10) + ' 条可衔接</div>')
        L("</div>")  # match-body
    L("</div>\n")  # match-card

L('</div></div>')

# Section 3: Summary table sorted by wait time
L('<div class="card"><div class="card-head">📌 排计划汇总建议表（按等待时间排序所有有效衔接）</div>')
L('<div style="overflow-x:auto"><table><thead><tr><th>#</th><th>干线船(白)</th><th>ETA PKG</th><th>布袋船(粉/橙)</th><th>ETD PKG</th><th>等待天</th><th>状态</th><th>布袋TEU</th><th>目的港</th><th>SVC</th></tr></thead><tbody>')

flat_matches=[]
for ci,c in enumerate(connections):
    for mm in c['matches']:
        flat_matches.append((c['main'],mm['f'],mm['delta'],mm['cls']))
flat_matches.sort(key=lambda x:x[2])

for i,(mw,mf,delta,cls) in enumerate(flat_matches,1):
    d=mf['dst']
    parts=[]
    for p in ['SOK','JED','MUN','NGB']:
        if d.get(p): parts.append(p+':'+d[p])
    dest_str=', '.join(parts) if parts else '-'
    status='✅快转' if cls=='good' else ('⚠️正常' if cls=='warn' else '🔴较慢')
    L('<tr><td>' + str(i) + '</td><td><strong>' + mw['vessel'] + '</strong></td><td class="date-f">' + mw['pkg'] + '</td>')
    L('<td><strong>' + mf['vessel'] + '</strong></td><td class="date-f">' + mf['pkg'] + '</td>')
    L('<td class="' + cls + '">' + str(delta) + '</td><td>' + status + '</td><td class="teu-n">' + str(mf['teu']) + '</td>')
    L('<td>' + dest_str + '</td><td>' + str(mf['svc']) + '</td></tr>')

L('</tbody></table></div></div>')
L('</body></html>')

H='\n'.join(lines)
out=r'c:\Users\leahliu\WorkBuddy\20260525175100\budai_plan.html'
with open(out,'w',encoding='utf-8') as f:
    f.write(H)

print("Done! White=%d, Pink=%d, Connections=%d" % (len(white), len(pink), total_conn))
print("Output:", out)
